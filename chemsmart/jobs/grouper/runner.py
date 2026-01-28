"""
Grouper job runners for molecular structure clustering.
This module implements runners that execute grouping jobs and contains
all grouper algorithm implementations. The base GrouperJobRunner provides
common functionality; specialized runners extend it with strategy-specific logic.
This module also contains all grouper classes migrated from utils/grouper.py.
"""

import logging
import multiprocessing
import os
import pickle
from abc import ABC, abstractmethod
from collections import defaultdict
from multiprocessing import RawArray, shared_memory
from multiprocessing.pool import ThreadPool
from typing import Iterable, List, Optional, Tuple

import networkx as nx
import numpy as np
import pandas as pd
from joblib import Parallel, delayed
from networkx.algorithms import isomorphism
from rdkit import Chem, DataStructs
from rdkit.Chem import TorsionFingerprints, rdMolDescriptors, rdMolHash
from rdkit.Chem.rdFingerprintGenerator import GetRDKitFPGenerator
from scipy.optimize import linear_sum_assignment
from scipy.spatial.distance import cdist

from chemsmart.io.molecules.structure import Molecule
from chemsmart.jobs.runner import JobRunner
from chemsmart.utils.periodictable import PeriodicTable
from chemsmart.utils.utils import kabsch_align

logger = logging.getLogger(__name__)


class GrouperJobRunner(JobRunner):
    """Job runner for molecular grouping/clustering jobs."""

    JOBTYPES = ["grouper"]
    PROGRAM = "grouper"
    FAKE = False
    SCRATCH = False

    # Strategy to grouper class mapping (populated after class definitions)
    GROUPER_CLASSES = {}

    def __init__(
        self, server, scratch=None, fake=False, scratch_dir=None, **kwargs
    ):
        if scratch is None:
            scratch = self.SCRATCH
        super().__init__(
            server=server,
            scratch=scratch,
            scratch_dir=scratch_dir,
            fake=fake,
            **kwargs,
        )

    def _prerun(self, job):
        if not os.path.exists(job.output_dir):
            os.makedirs(job.output_dir)

    def _get_command(self, job):
        return None

    def _get_executable(self):
        return None

    def _create_process(self, job, command, env):
        try:
            grouper = self._create_grouper(job)
            groups, group_indices = grouper.group()
            job._grouper = grouper
            job._groups = groups
            job._group_indices = group_indices
            self._write_outputs(job, groups, group_indices)
            return 0
        except Exception as e:
            logger.error(f"Error during grouping: {str(e)}")
            with open(job.errfile, "w") as err:
                err.write(f"Error during grouping: {str(e)}\n")
            return 1

    def _create_grouper(self, job):
        """Create appropriate grouper instance based on job strategy."""
        strategy = job.grouping_strategy

        if strategy not in self.GROUPER_CLASSES:
            raise ValueError(f"Unknown grouping strategy: {strategy}")

        grouper_cls = self.GROUPER_CLASSES[strategy]

        # Common kwargs
        kwargs = {
            "molecules": job.molecules,
            "num_procs": job.num_procs,
            "label": job.label,
        }

        # Strategy-specific kwargs
        if strategy in ["rmsd", "hrmsd", "spyrmsd", "irmsd", "pymolrmsd"]:
            kwargs["threshold"] = job.threshold
            kwargs["num_groups"] = job.num_groups
            kwargs["ignore_hydrogens"] = job.ignore_hydrogens
        elif strategy in ["tanimoto", "torsion"]:
            kwargs["threshold"] = job.threshold
            kwargs["num_groups"] = job.num_groups
        elif strategy == "connectivity":
            kwargs["threshold"] = job.threshold
        # isomorphism and formula don't need threshold

        # Add any additional kwargs from job
        kwargs.update(job.grouper_kwargs)

        return grouper_cls(**kwargs)

    def _write_outputs(self, job, groups, group_indices):
        """
        Write grouping results following the same logic as utils/grouper.py unique() method.

        - Creates per-group xyz files with molecules sorted by energy
        - Appends Groups summary to existing RMSD matrix excel file
        """
        import glob

        from openpyxl import load_workbook

        unique_molecules = []

        # Determine file prefix
        file_prefix = f"{job.label}_group" if job.label else "group"

        for i, (group, indices) in enumerate(zip(groups, group_indices)):
            # Create tuples of (molecule, original_index) for tracking
            mol_index_pairs = list(zip(group, indices))

            # Filter molecules that have energy information and sort by energy
            molecules_with_energy = [
                (mol, idx)
                for mol, idx in mol_index_pairs
                if mol.energy is not None
            ]
            molecules_without_energy = [
                (mol, idx)
                for mol, idx in mol_index_pairs
                if mol.energy is None
            ]

            # Sort molecules with energy by energy (ascending - lowest first)
            if molecules_with_energy:
                sorted_pairs = sorted(
                    molecules_with_energy, key=lambda pair: pair[0].energy
                )
                # Add molecules without energy at the end
                sorted_pairs.extend(molecules_without_energy)
            else:
                # If no molecules have energy, use original group order
                sorted_pairs = mol_index_pairs

            # Write group XYZ file with all molecules sorted by energy
            group_filename = os.path.join(
                job.output_dir, f"{file_prefix}_{i+1}.xyz"
            )
            with open(group_filename, "w") as f:
                for j, (mol, original_idx) in enumerate(sorted_pairs):
                    # Write the molecule coordinates
                    f.write(f"{mol.num_atoms}\n")

                    # Create comment line with energy info and original molecule index
                    if mol.energy is not None:
                        comment = f"Group {i+1} Molecule {j+1} Original_Index: {original_idx+1} Energy(Hartree): {mol.energy:.8f}"
                    else:
                        comment = f"Group {i+1} Molecule {j+1} Original_Index: {original_idx+1} Energy: N/A"

                    f.write(f"{comment}\n")

                    # Write coordinates
                    for symbol, position in zip(
                        mol.chemical_symbols, mol.positions
                    ):
                        f.write(
                            f"{symbol:2s} {position[0]:15.10f} {position[1]:15.10f} {position[2]:15.10f}\n"
                        )

            logger.info(
                f"Written group {i+1} with {len(sorted_pairs)} molecules to {group_filename}"
            )

            # Add the lowest energy molecule (first in sorted pairs) as representative
            unique_molecules.append(sorted_pairs[0][0])

        # Find the matrix excel file and append Groups sheet
        # New filename format: {label}_{GrouperClass}_T{threshold}.xlsx or {label}_{GrouperClass}_N{num_groups}.xlsx
        label_prefix = f"{job.label}_" if job.label else ""
        matrix_files = glob.glob(
            os.path.join(job.output_dir, f"{label_prefix}*_T*.xlsx")
        ) + glob.glob(os.path.join(job.output_dir, f"{label_prefix}*_N*.xlsx"))
        if matrix_files:
            matrix_excel_file = matrix_files[0]  # Use the first one found

            # Load existing workbook and add Groups sheet
            wb = load_workbook(matrix_excel_file)

            # Remove existing Groups sheet if present
            if "Groups" in wb.sheetnames:
                del wb["Groups"]

            # Create Groups sheet
            ws = wb.create_sheet("Groups")

            # Write header
            ws["A1"] = "Group"
            ws["B1"] = "Members"
            ws["C1"] = "Indices"

            # Write data
            for i, indices in enumerate(group_indices):
                ws[f"A{i+2}"] = i + 1
                ws[f"B{i+2}"] = len(indices)
                ws[f"C{i+2}"] = str([idx + 1 for idx in indices])

            # Auto-adjust column widths
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 10
            ws.column_dimensions["C"].width = 50

            wb.save(matrix_excel_file)
            logger.info(f"Added Groups sheet to {matrix_excel_file}")

        logger.info(
            f"Generated {len(groups)} group XYZ files in {job.output_dir}"
        )

    def _run(self, process, **kwargs):
        pass

    def _postrun(self, job):
        pass


def to_graph_wrapper(
    mol: Molecule, bond_cutoff_buffer: float = 0.0, adjust_H: bool = True
) -> nx.Graph:
    """
    Global helper function to call Molecule.to_graph() for multiprocessing.

    Provides a picklable wrapper for the Molecule.to_graph() method that
    can be used with multiprocessing pools.

    Args:
        mol (Molecule): Molecule instance to convert to graph.
        bond_cutoff_buffer (float): Buffer for bond cutoff distance.
        adjust_H (bool): Whether to adjust hydrogen bond detection.

    Returns:
        networkx.Graph: Molecular graph representation.
    """
    return mol.to_graph(
        bond_cutoff_buffer=bond_cutoff_buffer, adjust_H=adjust_H
    )


class StructureGrouperConfig:
    """
    Configuration container for StructureMatcher parameters.

    Stores tolerance parameters for structure matching algorithms.
    Default values are optimized for heterogeneous molecular systems
    and may need adjustment for specific molecular types.

    Attributes:
        ltol (float): Length tolerance for structure matching.
        stol (float): Site tolerance for atomic position matching.
        angle_tol (float): Angle tolerance in degrees for structure matching.
    """

    def __init__(self, ltol=0.1, stol=0.18, angle_tol=1):
        """
        Initialize structure grouper configuration.

        Args:
            ltol (float): Length tolerance. Defaults to 0.1.
            stol (float): Site tolerance. Defaults to 0.18.
            angle_tol (float): Angle tolerance in degrees. Defaults to 1.
        """
        self.ltol = ltol
        self.stol = stol
        self.angle_tol = angle_tol


class MoleculeGrouper(ABC):
    """
    Abstract base class for molecular structure grouping algorithms.

    Defines the common interface that all molecular grouping strategies
    must implement. Cannot be directly instantiated and designed to
    ensure consistent behavior across different grouping methods.

    Attributes:
        molecules (Iterable[Molecule]): Collection of molecules to group.
        num_procs (int): Number of processes for parallel computation.
        label (str): Label/name for this grouping task (used in output filenames).
    """

    def __init__(
        self,
        molecules: Iterable[Molecule],
        num_procs: int = 1,
        label: str = None,
    ):
        """
        Initialize the molecular grouper.

        Args:
            molecules (Iterable[Molecule]): Collection of molecules to group.
            num_procs (int): Number of processes for parallel computation.
                Defaults to 1.
            label (str): Label/name for this grouping task. Used in output folder
                and file names. Defaults to None.
        """
        self.molecules = molecules
        self.num_procs = int(max(1, num_procs))
        self.label = label

        # Cache for avoiding repeated grouping calculations
        self._cached_groups = None
        self._cached_group_indices = None

        self._validate_inputs()

    def _validate_inputs(self) -> None:
        """
        Validate input molecules for grouping.

        Ensures that the input is an iterable collection and all items
        are valid Molecule instances.

        Raises:
            TypeError: If molecules is not iterable or contains non-Molecule items.
        """
        if not isinstance(self.molecules, Iterable):
            raise TypeError("Molecules must be an iterable collection")
        if not all(isinstance(m, Molecule) for m in self.molecules):
            raise TypeError("All items must be Molecule instances")

    @abstractmethod
    def group(self) -> Tuple[List[List[Molecule]], List[List[int]]]:
        """
        Main grouping method to return grouped molecules and their indices.

        Must be implemented by subclasses to define specific grouping logic.

        Returns:
            Tuple[List[List[Molecule]], List[List[int]]]: Tuple containing:
                - List of molecule groups (each group is a list of molecules)
                - List of index groups (corresponding indices for each group)
        """
        pass

    def unique(
        self, output_dir: str = ".", prefix: str = "group"
    ) -> List[Molecule]:
        """
        Get unique representative molecules from each group.

        Returns the lowest energy molecule from each group as a representative
        of that structural family. Also generates XYZ files for each group,
        sorted by energy, in a dedicated subfolder.

        Args:
            output_dir (str): Base directory for output. Default is current directory.
            prefix (str): Prefix for output XYZ files. Default is "group".

        Returns:
            List[Molecule]: List of unique representative molecules (lowest energy from each group).
        """
        import os

        # Use cached results if available, otherwise compute and cache
        if (
            self._cached_groups is not None
            and self._cached_group_indices is not None
        ):
            print(f"[{self.__class__.__name__}] Using cached grouping results")
            groups, group_indices = (
                self._cached_groups,
                self._cached_group_indices,
            )
        else:
            print(
                f"[{self.__class__.__name__}] Computing groups for unique method"
            )
            groups, group_indices = self.group()
            # Cache the results
            self._cached_groups = groups
            self._cached_group_indices = group_indices

        unique_molecules = []

        # Create dedicated subfolder for XYZ files (include label if provided)
        if self.label:
            result_folder = f"{self.label}_group_result"
        else:
            result_folder = "group_result"
        full_output_path = os.path.join(output_dir, result_folder)
        os.makedirs(full_output_path, exist_ok=True)

        logger.info(f"Creating XYZ files in folder: {full_output_path}")

        # Determine the file prefix (include label if provided)
        if self.label:
            file_prefix = f"{self.label}_{prefix}"
        else:
            file_prefix = prefix

        for i, (group, indices) in enumerate(zip(groups, group_indices)):
            # Create tuples of (molecule, original_index) for tracking
            mol_index_pairs = list(zip(group, indices))

            # Filter molecules that have energy information and sort by energy
            molecules_with_energy = [
                (mol, idx)
                for mol, idx in mol_index_pairs
                if mol.energy is not None
            ]
            molecules_without_energy = [
                (mol, idx)
                for mol, idx in mol_index_pairs
                if mol.energy is None
            ]

            # Sort molecules with energy by energy (ascending - lowest first)
            if molecules_with_energy:
                sorted_pairs = sorted(
                    molecules_with_energy, key=lambda pair: pair[0].energy
                )
                # Add molecules without energy at the end
                sorted_pairs.extend(molecules_without_energy)
            else:
                # If no molecules have energy, use original group order
                sorted_pairs = mol_index_pairs

            # Write group XYZ file with all molecules sorted by energy
            group_filename = os.path.join(
                full_output_path, f"{file_prefix}_{i+1}.xyz"
            )
            with open(group_filename, "w") as f:
                for j, (mol, original_idx) in enumerate(sorted_pairs):
                    # Write the molecule coordinates
                    f.write(f"{mol.num_atoms}\n")

                    # Create comment line with energy info and original molecule index
                    if mol.energy is not None:
                        comment = f"Group {i+1} Member {j+1} Original_Index: {original_idx+1} Energy(Hartree): {mol.energy:.8f}"
                    else:
                        comment = f"Group {i+1} Member {j+1} Original_Index: {original_idx+1} Energy: N/A"

                    f.write(f"{comment}\n")

                    # Write coordinates
                    for symbol, position in zip(
                        mol.chemical_symbols, mol.positions
                    ):
                        f.write(
                            f"{symbol:2s} {position[0]:15.10f} {position[1]:15.10f} {position[2]:15.10f}\n"
                        )

            logger.info(
                f"Written group {i+1} with {len(sorted_pairs)} molecules to {group_filename}"
            )

            # Add the lowest energy molecule (first in sorted pairs) as representative
            unique_molecules.append(sorted_pairs[0][0])

        logger.info(
            f"Generated {len(groups)} group XYZ files in {full_output_path}"
        )

        return unique_molecules


class RMSDGrouper(MoleculeGrouper):
    """
    Abstract base class for RMSD-based molecular grouping.

    Groups molecules based on geometric similarity of atomic positions using
    various RMSD calculation methods. This base class provides common
    functionality for all RMSD-based groupers while allowing subclasses
    to implement specific RMSD calculation algorithms.

    Follows the same design pattern as JobRunner - provides the core
    grouping logic while allowing subclasses to implement specific
    RMSD calculation methods via _calculate_rmsd_core().

    Attributes:
        molecules (Iterable[Molecule]): Inherited; collection of molecules to
            group.
        num_procs (int): Inherited; number of worker processes/threads.
        threshold (float): RMSD threshold for grouping molecules.
        align_molecules (bool): Whether to align molecules before RMSD calculation.
        ignore_hydrogens (bool): Whether to exclude hydrogen atoms from RMSD.
    """

    def __init__(
        self,
        molecules: Iterable[Molecule],
        threshold=None,  # RMSD threshold for grouping
        num_groups=None,  # Number of groups to create (alternative to threshold)
        num_procs: int = 1,
        align_molecules: bool = True,
        ignore_hydrogens: bool = False,
        label: str = None,  # Label for output files
        **kwargs,  # Option to ignore H atoms for grouping
    ):
        """
        Initialize RMSD-based molecular grouper.

        Args:
            molecules (Iterable[Molecule]): Collection of molecules to group.
            threshold (float): RMSD threshold for grouping. Defaults to 0.5.
                Ignored if num_groups is specified.
            num_groups (int): Number of groups to create. When specified,
                automatically determines threshold to create this many groups.
            num_procs (int): Number of processes for parallel computation.
            align_molecules (bool): Whether to align molecules using Kabsch
                algorithm before RMSD calculation. Defaults to True.
            ignore_hydrogens (bool): Whether to exclude hydrogen atoms from
                RMSD calculation. Defaults to False.
            label (str): Label/name for output files. Defaults to None.

        Note:
            Uses complete linkage clustering: a structure joins a group only if
            its RMSD to ALL existing members is below the threshold. This prevents
            the chaining effect where dissimilar structures end up in the same
            group through intermediate "bridge" structures.
        """
        super().__init__(molecules, num_procs, label=label)

        # Validate that threshold and num_groups are mutually exclusive
        if threshold is not None and num_groups is not None:
            raise ValueError(
                "Cannot specify both threshold (-t) and num_groups (-N). Please use only one."
            )

        if threshold is None and num_groups is None:
            threshold = 0.5
        self.threshold = threshold  # RMSD threshold for grouping
        self.num_groups = num_groups  # Number of groups to create
        self.align_molecules = align_molecules
        self.ignore_hydrogens = ignore_hydrogens
        # Cache sorted chemical symbols as sets for faster comparison
        self._chemical_symbol_sets = [
            set(mol.chemical_symbols) for mol in molecules
        ]

    def _get_heavy_atoms(self, mol: Molecule) -> Tuple[np.ndarray, List[str]]:
        """
        Extract heavy atoms (non-hydrogen) if ignore_hydrogens is enabled.

        Args:
            mol (Molecule): Molecule to process.

        Returns:
            Tuple[np.ndarray, List[str]]: Tuple containing positions array
                and chemical symbols list (filtered or full based on settings).
        """
        if self.ignore_hydrogens:
            non_h_indices = [
                i for i, sym in enumerate(mol.chemical_symbols) if sym != "H"
            ]
            return mol.positions[non_h_indices], [
                mol.chemical_symbols[i] for i in non_h_indices
            ]
        return (
            mol.positions,
            mol.chemical_symbols,
        )  # Use all atoms if flag is False

    @abstractmethod
    def _calculate_rmsd(self, idx_pair: Tuple[int, int]) -> float:
        """Calculate RMSD between two molecules. Must be implemented by subclasses."""
        pass

    def group(self) -> Tuple[List[List[Molecule]], List[List[int]]]:
        """
        Group molecules by geometric similarity using RMSD clustering.

        Computes pairwise RMSD values between all molecules and groups
        those within the specified threshold using connected components
        clustering, or automatically determines threshold to create
        the specified number of groups. Automatically saves RMSD matrix
        to group_result folder.

        Returns:
            Tuple[List[List[Molecule]], List[List[int]]]: Tuple containing:
                - List of molecule groups (each group is a list of molecules)
                - List of index groups (corresponding indices for each group)
        """
        import time

        # Record start time for grouping process
        grouping_start_time = time.time()

        n = len(self.molecules)
        indices = [(i, j) for i in range(n) for j in range(i + 1, n)]
        total_pairs = len(indices)

        print(
            f"[{self.__class__.__name__}] Starting calculation for {n} molecules ({total_pairs} pairs)"
        )

        # For real-time output, calculate one by one instead of using multiprocessing
        rmsd_values = []
        for idx, (i, j) in enumerate(indices):
            rmsd = self._calculate_rmsd((i, j))
            rmsd_values.append(rmsd)
            print(
                f"The {idx+1}/{total_pairs} pair (conformer{i+1}, conformer{j+1}) calculation finished, RMSD= {rmsd:.7f}"
            )

        # Build full RMSD matrix for output
        rmsd_matrix = np.zeros((n, n))
        for (i, j), rmsd in zip(indices, rmsd_values):
            rmsd_matrix[i, j] = rmsd_matrix[j, i] = rmsd

        # Save RMSD matrix to group_result folder
        import os

        # Use label for folder name if provided
        if self.label:
            output_dir = f"{self.label}_group_result"
        else:
            output_dir = "group_result"
        os.makedirs(output_dir, exist_ok=True)

        # Create filename based on label, grouper type and threshold/num_groups
        # Format: {label}_{GrouperClass}_T{threshold}.xlsx or {label}_{GrouperClass}_N{num_groups}.xlsx
        label_prefix = f"{self.label}_" if self.label else ""
        if self.num_groups is not None:
            matrix_filename = os.path.join(
                output_dir,
                f"{label_prefix}{self.__class__.__name__}_N{self.num_groups}.xlsx",
            )
        else:
            matrix_filename = os.path.join(
                output_dir,
                f"{label_prefix}{self.__class__.__name__}_T{self.threshold}.xlsx",
            )

        # Choose grouping strategy based on parameters (do this first to set _auto_threshold)
        if self.num_groups is not None:
            groups, index_groups = self._group_by_num_groups(
                rmsd_matrix, rmsd_values, indices
            )
        else:
            groups, index_groups = self._group_by_threshold(
                rmsd_values, indices
            )

        # Calculate total grouping time
        grouping_end_time = time.time()
        grouping_time = grouping_end_time - grouping_start_time

        # Save full matrix (after grouping to include auto-determined threshold)
        self._save_rmsd_matrix(rmsd_matrix, matrix_filename, grouping_time)

        # Cache the results to avoid recomputation
        self._cached_groups = groups
        self._cached_group_indices = index_groups

        return groups, index_groups

    def _group_by_threshold(self, rmsd_values, indices):
        """
        Threshold-based grouping using complete linkage.

        A structure joins a group only if its RMSD to ALL existing members
        is below the threshold. This prevents the chaining effect.
        """
        n = len(self.molecules)

        # Build adjacency matrix for clustering
        adj_matrix = np.zeros((n, n), dtype=bool)
        for (i, j), rmsd in zip(indices, rmsd_values):
            if rmsd < self.threshold:
                adj_matrix[i, j] = adj_matrix[j, i] = True

        # Complete linkage grouping
        groups, index_groups = self._complete_linkage_grouping(adj_matrix, n)

        return groups, index_groups

    def _complete_linkage_grouping(self, adj_matrix, n):
        """
        Perform complete linkage grouping (from last to first).

        A structure joins a group only if its RMSD to ALL existing members
        of that group is below the threshold (i.e., all edges exist in adj_matrix).
        This prevents the chaining effect where dissimilar structures end up
        in the same group through intermediate "bridge" structures.

        Iterates from last to first structure so that higher-energy structures
        (typically at the end) are more likely to form singleton groups,
        preserving lower-energy structures in larger groups.

        Args:
            adj_matrix: Boolean adjacency matrix where adj_matrix[i,j] = True
                       if RMSD between i and j is below threshold
            n: Number of molecules

        Returns:
            Tuple of (groups, index_groups)
        """
        assigned = [False] * n
        groups = []
        index_groups = []

        # Iterate from last to first (higher energy structures first as seeds)
        for i in range(n - 1, -1, -1):
            if assigned[i]:
                continue

            # Start a new group with molecule i
            current_group = [i]
            assigned[i] = True

            # Try to add unassigned molecules with lower indices (lower energy)
            for j in range(i - 1, -1, -1):
                if assigned[j]:
                    continue

                # Check if j is connected to ALL members in current_group
                can_join = all(
                    adj_matrix[j, member] for member in current_group
                )

                if can_join:
                    current_group.append(j)
                    assigned[j] = True

            # Sort indices within group (lowest index = lowest energy first)
            current_group.sort()
            # Add the completed group
            groups.append([self.molecules[idx] for idx in current_group])
            index_groups.append(current_group)

        # Reverse groups so that groups containing lower-energy structures come first
        groups.reverse()
        index_groups.reverse()

        return groups, index_groups

    def _group_by_num_groups(self, rmsd_matrix, rmsd_values, indices):
        """Automatic grouping to create specified number of groups."""
        n = len(self.molecules)

        if self.num_groups >= n:
            # If requesting more groups than molecules, each molecule is its own group
            print(
                f"[{self.__class__.__name__}] Requested {self.num_groups} groups but only {n} molecules. Creating {n} groups."
            )
            groups = [[mol] for mol in self.molecules]
            index_groups = [[i] for i in range(n)]
            return groups, index_groups

        # Find appropriate threshold to create desired number of groups
        threshold = self._find_optimal_threshold(rmsd_values, indices, n)

        # Store the auto-determined threshold for summary reporting
        self._auto_threshold = threshold

        print(
            f"[{self.__class__.__name__}] Auto-determined threshold: {threshold:.7f} to create {self.num_groups} groups"
        )

        # Build adjacency matrix with the determined threshold
        adj_matrix = np.zeros((n, n), dtype=bool)
        for (i, j), rmsd in zip(indices, rmsd_values):
            if rmsd < threshold:
                adj_matrix[i, j] = adj_matrix[j, i] = True

        # Use complete linkage grouping
        groups, index_groups = self._complete_linkage_grouping(adj_matrix, n)
        actual_groups = len(groups)

        print(
            f"[{self.__class__.__name__}] Created {actual_groups} groups (requested: {self.num_groups})"
        )

        if actual_groups > self.num_groups:
            # If we have too many groups, merge the smallest ones
            groups, index_groups = self._merge_groups_to_target(
                groups, index_groups, adj_matrix
            )

        return groups, index_groups

    def _find_optimal_threshold(self, rmsd_values, indices, n):
        """Find threshold that creates approximately the desired number of groups using binary search."""
        # Sort RMSD values to try different thresholds
        sorted_rmsd = sorted(
            [rmsd for rmsd in rmsd_values if not np.isinf(rmsd)]
        )

        if not sorted_rmsd:
            return 0.0

        # Binary search for optimal threshold
        low, high = 0, len(sorted_rmsd) - 1
        best_threshold = sorted_rmsd[-1]

        while low <= high:
            mid = (low + high) // 2
            threshold = sorted_rmsd[mid]

            # Build adjacency matrix with this threshold
            adj_matrix = np.zeros((n, n), dtype=bool)
            for (idx_i, idx_j), rmsd in zip(indices, rmsd_values):
                if rmsd < threshold:
                    adj_matrix[idx_i, idx_j] = adj_matrix[idx_j, idx_i] = True

            # Use complete linkage to count groups
            groups, _ = self._complete_linkage_grouping(adj_matrix, n)
            num_groups_found = len(groups)

            if num_groups_found == self.num_groups:
                # Found exact match
                return threshold
            elif num_groups_found > self.num_groups:
                # Too many groups, need higher threshold (more permissive)
                low = mid + 1
            else:
                # Too few groups, need lower threshold (more restrictive)
                best_threshold = threshold
                high = mid - 1

        return best_threshold

    def _merge_smallest_groups(self, labels, actual_groups):
        """Merge smallest groups to reach target number of groups."""
        # This is a simplified approach - in practice you might want more sophisticated merging
        unique_labels = np.unique(labels)
        group_sizes = [
            (label, len(np.where(labels == label)[0]))
            for label in unique_labels
        ]
        group_sizes.sort(key=lambda x: x[1])  # Sort by size

        # Keep largest groups, merge smallest ones into the largest
        groups_to_keep = group_sizes[-self.num_groups :]
        groups_to_merge = group_sizes[: -self.num_groups]

        # Create final groups
        groups = []
        index_groups = []

        # Add the groups we're keeping
        for label, _ in groups_to_keep:
            indices = list(np.where(labels == label)[0])
            groups.append([self.molecules[i] for i in indices])
            index_groups.append(indices)

        # Merge remaining groups into the largest group
        if groups_to_merge:
            merge_indices = []
            for label, _ in groups_to_merge:
                merge_indices.extend(list(np.where(labels == label)[0]))

            # Add to the largest group
            if groups:
                groups[0].extend([self.molecules[i] for i in merge_indices])
                index_groups[0].extend(merge_indices)

        return groups, index_groups

    def _merge_groups_to_target(self, groups, index_groups, adj_matrix):
        """
        Merge groups to reach target number when using complete linkage.

        Merges smallest groups into the most compatible larger groups,
        where compatibility is determined by the number of edges in adj_matrix.

        Args:
            groups: List of molecule groups
            index_groups: List of index groups
            adj_matrix: Boolean adjacency matrix

        Returns:
            Tuple of (merged_groups, merged_index_groups)
        """
        while len(groups) > self.num_groups:
            # Find the smallest group
            min_size = float("inf")
            min_idx = 0
            for i, g in enumerate(groups):
                if len(g) < min_size:
                    min_size = len(g)
                    min_idx = i

            # Find the best group to merge with (most connections)
            best_merge_idx = -1
            best_connection_count = -1

            for i, target_indices in enumerate(index_groups):
                if i == min_idx:
                    continue

                # Count connections between source group and target group
                connection_count = 0
                for src_idx in index_groups[min_idx]:
                    for tgt_idx in target_indices:
                        if adj_matrix[src_idx, tgt_idx]:
                            connection_count += 1

                if connection_count > best_connection_count:
                    best_connection_count = connection_count
                    best_merge_idx = i

            # Merge the smallest group into the best target
            if best_merge_idx >= 0:
                groups[best_merge_idx].extend(groups[min_idx])
                index_groups[best_merge_idx].extend(index_groups[min_idx])
            else:
                # No connections found, merge into the largest group
                largest_idx = max(
                    range(len(groups)),
                    key=lambda i: len(groups[i]) if i != min_idx else -1,
                )
                groups[largest_idx].extend(groups[min_idx])
                index_groups[largest_idx].extend(index_groups[min_idx])

            # Remove the merged group
            groups.pop(min_idx)
            index_groups.pop(min_idx)

        return groups, index_groups

    def __repr__(self):
        if self.num_groups is not None:
            return (
                f"{self.__class__.__name__}(num_groups={self.num_groups}, "
                f"num_procs={self.num_procs}, align_molecules={self.align_molecules}, "
                f"ignore_hydrogens={self.ignore_hydrogens})"
            )
        else:
            return (
                f"{self.__class__.__name__}(threshold={self.threshold}, "
                f"num_procs={self.num_procs}, align_molecules={self.align_molecules}, "
                f"ignore_hydrogens={self.ignore_hydrogens})"
            )

    def calculate_full_rmsd_matrix(
        self, output_file: Optional[str] = None
    ) -> np.ndarray:
        """
        Calculate the full RMSD matrix for all molecule pairs.

        Args:
            output_file (str, optional): Path to save RMSD matrix as text file

        Returns:
            np.ndarray: Symmetric RMSD matrix (n x n)
        """
        n = len(self.molecules)
        rmsd_matrix = np.zeros((n, n))

        logger.info(f"Calculating full RMSD matrix for {n} molecules")

        # Calculate upper triangular matrix (symmetric)
        indices = [(i, j) for i in range(n) for j in range(i + 1, n)]

        # Use multiprocessing for efficiency
        with multiprocessing.Pool(self.num_procs) as pool:
            rmsd_values = pool.map(self._calculate_rmsd, indices)

        # Fill the matrix
        for (i, j), rmsd in zip(indices, rmsd_values):
            rmsd_matrix[i, j] = rmsd_matrix[j, i] = rmsd

        # Output results to console
        col_width = 12  # Fixed width: fits "-XX.XXXXXX" (6 decimals)
        row_header_width = 7  # Width for row headers

        print(f"\nFull RMSD Matrix ({n}x{n}):")
        print("=" * (row_header_width + col_width * n))

        # Print column index header (right-aligned to match matrix values)
        header_line = " " * row_header_width
        for j in range(n):
            header_line += f"{j+1:>{col_width}}"
        print(header_line)

        # Print header row label with separator
        separator_line = "Conf".rjust(row_header_width)
        for j in range(n):
            separator_line += "-" * col_width
        print(separator_line)

        # Print matrix rows (right-aligned values)
        for i in range(n):
            row_line = f"{i+1:>{row_header_width}}"
            for j in range(n):
                if np.isinf(rmsd_matrix[i, j]):
                    row_line += f"{'∞':>{col_width}}"
                else:
                    row_line += f"{rmsd_matrix[i, j]:>{col_width}.6f}"
            print(row_line)

        # Save to file if requested
        if output_file:
            import os

            # Create directory if it doesn't exist
            os.makedirs(
                (
                    os.path.dirname(output_file)
                    if os.path.dirname(output_file)
                    else "."
                ),
                exist_ok=True,
            )

            with open(output_file, "w") as f:
                f.write(
                    f"Full RMSD Matrix ({n}x{n}) - {self.__class__.__name__}\n"
                )
                f.write("=" * (row_header_width + col_width * n) + "\n")
                f.write("Values in Angstroms (Å)\n")
                f.write("∞ indicates non-comparable molecules\n")
                f.write("-" * (row_header_width + col_width * n) + "\n\n")

                # Write column index header (right-aligned to match matrix values)
                header_line = " " * row_header_width
                for j in range(n):
                    header_line += f"{j+1:>{col_width}}"
                f.write(header_line + "\n")

                # Write header row label with separator line
                separator_line = "Conf".rjust(row_header_width)
                for j in range(n):
                    separator_line += "-" * col_width
                f.write(separator_line + "\n")

                # Write matrix with 6 decimal places (right-aligned)
                for i in range(n):
                    row_line = f"{i+1:>{row_header_width}}"
                    for j in range(n):
                        if np.isinf(rmsd_matrix[i, j]):
                            row_line += f"{'∞':>{col_width}}"
                        else:
                            row_line += f"{rmsd_matrix[i, j]:>{col_width}.6f}"
                    f.write(row_line + "\n")

            logger.info(f"RMSD matrix saved to {output_file}")

        return rmsd_matrix

    def _save_rmsd_matrix(
        self,
        rmsd_matrix: np.ndarray,
        filename: str,
        grouping_time: float = None,
    ):
        """Save RMSD matrix to Excel file with 7 decimal precision."""
        n = rmsd_matrix.shape[0]

        # Change extension to .xlsx if it's .txt
        if filename.endswith(".txt"):
            filename = filename[:-4] + ".xlsx"
        elif not filename.endswith(".xlsx"):
            filename = filename + ".xlsx"

        # Create DataFrame with simple numeric indices as row and column labels
        row_labels = [str(i + 1) for i in range(n)]
        col_labels = [str(j + 1) for j in range(n)]

        # Replace inf with string "∞" for display
        matrix_display = np.where(np.isinf(rmsd_matrix), np.nan, rmsd_matrix)
        df = pd.DataFrame(matrix_display, index=row_labels, columns=col_labels)

        # Create Excel writer with openpyxl engine
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            # Write matrix to 'RMSD_Matrix' sheet starting from row 8 to leave room for header info
            df.to_excel(
                writer,
                sheet_name="RMSD_Matrix",
                startrow=8,
                float_format="%.7f",
            )

            # Get the worksheet to add header information
            worksheet = writer.sheets["RMSD_Matrix"]

            # Add header information
            row = 1
            worksheet[f"A{row}"] = (
                f"Full RMSD Matrix ({n}x{n}) - {self.__class__.__name__}"
            )
            row += 1

            # Threshold or num_groups
            if hasattr(self, "num_groups") and self.num_groups is not None:
                worksheet[f"A{row}"] = (
                    f"Requested Groups (-N): {self.num_groups}"
                )
                row += 1
                if (
                    hasattr(self, "_auto_threshold")
                    and self._auto_threshold is not None
                ):
                    worksheet[f"A{row}"] = (
                        f"Auto-determined Threshold: {self._auto_threshold:.7f} Å"
                    )
                    row += 1
            else:
                worksheet[f"A{row}"] = f"Threshold: {self.threshold} Å"
                row += 1

            # Parameters specific to different groupers
            worksheet[f"A{row}"] = (
                f"Align Molecules: {getattr(self, 'align_molecules', 'N/A')}"
            )
            row += 1
            worksheet[f"A{row}"] = (
                f"Ignore Hydrogens: {getattr(self, 'ignore_hydrogens', False)}"
            )
            row += 1

            # IRMSDGrouper specific parameters
            if hasattr(self, "check_stereo"):
                worksheet[f"A{row}"] = f"Check Stereo: {self.check_stereo}"
                row += 1

            if grouping_time is not None:
                worksheet[f"A{row}"] = (
                    f"Grouping Time: {grouping_time:.2f} seconds"
                )
                row += 1

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = min(max_length + 2, 18)
                worksheet.column_dimensions[column_letter].width = (
                    adjusted_width
                )

        logger.info(f"RMSD matrix saved to {filename}")


class BasicRMSDGrouper(RMSDGrouper):
    """
    Basic RMSD grouper using standard Euclidean distance calculation.

    Implements the most straightforward RMSD calculation using the standard
    formula: sqrt(mean(sum((pos1 - pos2)^2))). This is the classic RMSD
    implementation that compares atomic positions directly.
    """

    def _calculate_rmsd(self, idx_pair: Tuple[int, int]) -> float:
        """Calculate RMSD between two molecules without atom reordering."""
        i, j = idx_pair
        mol1, mol2 = self.molecules[i], self.molecules[j]

        if self.ignore_hydrogens:
            pos1, symbols1 = self._get_heavy_atoms(mol1)
            pos2, symbols2 = self._get_heavy_atoms(mol2)
        else:
            pos1, symbols1 = mol1.positions, list(mol1.chemical_symbols)
            pos2, symbols2 = mol2.positions, list(mol2.chemical_symbols)

        # Quick check for compatibility
        if len(symbols1) != len(symbols2) or sorted(symbols1) != sorted(
            symbols2
        ):
            return np.inf

        if self.align_molecules:
            logger.debug("Aligning molecules using Kabsch algorithm.")
            pos1, pos2, _, _, _ = kabsch_align(pos1, pos2)
        rmsd = np.sqrt(np.mean(np.sum((pos1 - pos2) ** 2, axis=1)))
        return rmsd


class HungarianRMSDGrouper(RMSDGrouper):
    """
    Hungarian RMSD grouper for optimal atom assignment.

    Uses the Hungarian algorithm (Kuhn-Munkres algorithm) to find the optimal
    assignment of atoms between two molecules that minimizes RMSD. This approach
    handles cases where atom ordering might differ between molecules of the same
    chemical structure.

    The Hungarian algorithm ensures that atoms of the same element type are
    optimally paired to minimize the sum of squared distances, resulting in
    more accurate RMSD calculations for molecules with permuted atom arrangements.
    """

    def __init__(
        self,
        molecules,
        threshold=None,
        num_groups=None,
        num_procs: int = 1,
        align_molecules: bool = True,
        ignore_hydrogens: bool = False,
        label: str = None,
        **kwargs,
    ):
        """
        Initialize Hungarian RMSD grouper.

        Args:
            molecules: Collection of molecules to group.
            threshold (float): RMSD threshold for grouping.
            num_groups (int): Number of groups to create (alternative to threshold).
            num_procs (int): Number of processes for parallel computation.
            align_molecules (bool): Whether to align molecules (legacy parameter).
            ignore_hydrogens (bool): Whether to exclude hydrogen atoms.
            label (str): Label for output files.
        """
        super().__init__(
            molecules,
            threshold,
            num_groups,
            num_procs,
            align_molecules,
            ignore_hydrogens,
            label=label,
        )

    def _calculate_rmsd(self, idx_pair):
        i, j = idx_pair
        mol1, mol2 = self.molecules[i], self.molecules[j]

        if self.ignore_hydrogens:
            pos1, symbols1 = self._get_heavy_atoms(mol1)
            pos2, symbols2 = self._get_heavy_atoms(mol2)
        else:
            pos1, symbols1 = mol1.positions, list(mol1.chemical_symbols)
            pos2, symbols2 = mol2.positions, list(mol2.chemical_symbols)

        if len(symbols1) != len(symbols2) or sorted(symbols1) != sorted(
            symbols2
        ):
            return np.inf

        # Use Hungarian algorithm for optimal atom matching
        elements = sorted(set(symbols1))
        matched_idx1 = []
        matched_idx2 = []
        for elem in elements:
            idxs1 = [k for k, s in enumerate(symbols1) if s == elem]
            idxs2 = [k for k, s in enumerate(symbols2) if s == elem]

            if len(idxs1) == 1 and len(idxs2) == 1:
                # For single atoms, direct match
                matched_idx1.extend(idxs1)
                matched_idx2.extend(idxs2)
            else:
                # For multiple atoms of same type, use Hungarian algorithm
                pos1_elem = pos1[idxs1]
                pos2_elem = pos2[idxs2]
                # Use scipy's cdist for efficient distance matrix calculation
                dist_matrix = cdist(pos1_elem, pos2_elem, metric="sqeuclidean")
                row_ind, col_ind = linear_sum_assignment(dist_matrix)
                matched_idx1.extend([idxs1[r] for r in row_ind])
                matched_idx2.extend([idxs2[c] for c in col_ind])

        pos1_matched = pos1[matched_idx1]
        pos2_matched = pos2[matched_idx2]

        if self.align_molecules:
            logger.debug("Aligning molecules using Kabsch algorithm.")
            pos1_matched, pos2_matched, _, _, rmsd = kabsch_align(
                pos1_matched, pos2_matched
            )

        rmsd = np.sqrt(
            np.mean(np.sum((pos1_matched - pos2_matched) ** 2, axis=1))
        )
        return rmsd


class SpyRMSDGrouper(RMSDGrouper):
    """
    SpyRMSD grouper with graph isomorphism symmetry correction.

    Uses graph isomorphism for advanced RMSD calculations with symmetry correction.
    This approach handles molecular symmetries and atom permutations more
    comprehensively than basic Hungarian assignment.

    Supports optimal superposition and can handle cases where molecular graphs
    need to be compared for finding the best atom mapping.
    """

    def __init__(
        self,
        molecules,
        threshold=None,
        num_groups=None,
        num_procs: int = 1,
        align_molecules: bool = True,
        ignore_hydrogens: bool = False,
        cache: bool = True,  # Cache graph isomorphisms
        label: str = None,
        **kwargs,
    ):
        """
        Initialize SpyRMSD-based molecular grouper.

        Args:
            molecules: Collection of molecules to group.
            threshold (float): RMSD threshold for grouping. Defaults to 0.5.
            num_groups (int): Number of groups to create (alternative to threshold).
            num_procs (int): Number of processes for parallel computation.
            align_molecules (bool): Whether to align molecules (legacy parameter).
            ignore_hydrogens (bool): Whether to exclude hydrogen atoms.
            cache (bool): Cache graph isomorphisms for efficiency.
            label (str): Label for output files.
        """
        super().__init__(
            molecules,
            threshold,
            num_groups,
            num_procs,
            align_molecules,
            ignore_hydrogens,
            label=label,
        )
        self.cache = cache
        self.periodic_table = PeriodicTable()
        # Store best isomorphisms for each molecule pair
        self.best_isomorphisms = {}

    @staticmethod
    def _center_coordinates(pos: np.ndarray) -> np.ndarray:
        """
        Calculate the center of geometry (centroid) of atomic coordinates.

        Args:
            pos (np.ndarray): Array of atomic coordinates (N x 3).

        Returns:
            np.ndarray: Center of geometry as a 3D vector.
        """
        return np.mean(pos, axis=0)

    def _symbol_to_atomicnum(self, symbol: str) -> int:
        """
        Convert element symbol to atomic number using PeriodicTable.

        Args:
            symbol (str): Element symbol (e.g., 'H', 'C', 'O').

        Returns:
            int: Atomic number, or 0 if unknown.
        """
        try:
            return self.periodic_table.to_atomic_number(symbol)
        except (ValueError, IndexError):
            logger.warning(f"Unknown element symbol: {symbol}")
            return 0

    def _symmrmsd(
        self,
        pos1: np.ndarray,
        pos2: np.ndarray,
        symbols1: list,
        symbols2: list,
        adj_matrix1: np.ndarray,
        adj_matrix2: np.ndarray,
        mol_idx_pair: Tuple[int, int] = None,
    ) -> float:
        """
        Calculate symmetry-corrected RMSD using graph isomorphism.

        Computes RMSD and internally stores the best isomorphism mapping.
        This provides essential chemical information while keeping the interface
        consistent with other RMSD methods.

        Args:
            pos1 (np.ndarray): First set of coordinates (N x 3).
            pos2 (np.ndarray): Second set of coordinates (N x 3).
            symbols1 (list): Chemical symbols for first structure.
            symbols2 (list): Chemical symbols for second structure.
            adj_matrix1 (np.ndarray): Adjacency matrix for first structure.
            adj_matrix2 (np.ndarray): Adjacency matrix for second structure.
            mol_idx_pair (Tuple[int, int], optional): Molecule pair indices for storing mapping.

        Returns:
            float: Symmetry-corrected RMSD value.

        Notes:
            Returns np.inf if graphs are not isomorphic, indicating
            fundamentally different molecular structures. Best isomorphism
            is stored internally for later retrieval.
        """

        # Verify same number of atoms
        if len(pos1) != len(pos2):
            if mol_idx_pair:
                self.best_isomorphisms[mol_idx_pair] = None
            return np.inf

        # Verify same atomic composition
        if sorted(symbols1) != sorted(symbols2):
            if mol_idx_pair:
                self.best_isomorphisms[mol_idx_pair] = None
            return np.inf

        # Create NetworkX graphs from adjacency matrices
        G1 = nx.from_numpy_array(adj_matrix1)
        G2 = nx.from_numpy_array(adj_matrix2)

        # Add atomic numbers as node attributes
        for i, symbol in enumerate(symbols1):
            G1.nodes[i]["element"] = self._symbol_to_atomicnum(symbol)
        for i, symbol in enumerate(symbols2):
            G2.nodes[i]["element"] = self._symbol_to_atomicnum(symbol)

        # Find all graph isomorphisms (atom mappings)
        node_match = isomorphism.categorical_node_match("element", 0)
        GM = isomorphism.GraphMatcher(G1, G2, node_match=node_match)

        if not GM.is_isomorphic():
            # If not isomorphic, store None and return infinity
            if mol_idx_pair:
                self.best_isomorphisms[mol_idx_pair] = None
            return np.inf

        # Get all isomorphisms at once
        all_isomorphisms = list(GM.isomorphisms_iter())

        # Find minimum RMSD among all isomorphisms
        min_rmsd = np.inf
        best_isomorphism = None

        for mapping in all_isomorphisms:
            # Create index arrays for this mapping
            idx1 = list(range(len(symbols1)))
            idx2 = [mapping[i] for i in range(len(symbols2))]

            # Reorder pos2 according to this mapping
            reordered_pos2 = np.array(
                [pos2[idx2[i]] for i in range(len(pos2))]
            )

            # Calculate RMSD for this mapping
            if self.align_molecules:
                _, _, _, _, rmsd = kabsch_align(pos1, reordered_pos2)
            else:
                diff = pos1 - reordered_pos2
                rmsd = np.sqrt(np.mean(np.sum(diff**2, axis=1)))

            if rmsd < min_rmsd:
                min_rmsd = rmsd
                best_isomorphism = (idx1, idx2)

        # Store the best isomorphism if molecule indices provided
        if mol_idx_pair:
            self.best_isomorphisms[mol_idx_pair] = best_isomorphism

        return min_rmsd

    def _calculate_rmsd(self, idx_pair):
        """Calculate RMSD using symmetry correction with graph isomorphism.

        Maintains compatibility with parent class by returning only RMSD value,
        while internally storing best isomorphism mappings for chemical analysis.
        """

        i, j = idx_pair
        mol1, mol2 = self.molecules[i], self.molecules[j]

        if self.ignore_hydrogens:
            pos1, symbols1 = self._get_heavy_atoms(mol1)
            pos2, symbols2 = self._get_heavy_atoms(mol2)
        else:
            pos1, symbols1 = mol1.positions, list(mol1.chemical_symbols)
            pos2, symbols2 = mol2.positions, list(mol2.chemical_symbols)

        # Quick compatibility check
        if len(symbols1) != len(symbols2) or sorted(symbols1) != sorted(
            symbols2
        ):
            # Store no mapping for incompatible molecules
            self.best_isomorphisms[(i, j)] = None
            return np.inf

        # Use symmetry-corrected RMSD if molecules have connectivity
        # Try to get molecular graphs for symmetry correction
        adj_matrix1 = (
            mol1.adjacency_matrix
            if hasattr(mol1, "adjacency_matrix")
            else None
        )
        adj_matrix2 = (
            mol2.adjacency_matrix
            if hasattr(mol2, "adjacency_matrix")
            else None
        )

        # If no adjacency matrices exist, try to generate them from molecular graphs
        if adj_matrix1 is None or adj_matrix2 is None:
            try:
                # Generate graph and adjacency matrix for mol1
                if adj_matrix1 is None:
                    graph1 = mol1.to_graph()
                    adj_matrix1 = nx.adjacency_matrix(graph1).toarray()

                # Generate graph and adjacency matrix for mol2
                if adj_matrix2 is None:
                    graph2 = mol2.to_graph()
                    adj_matrix2 = nx.adjacency_matrix(graph2).toarray()

                logger.debug(
                    "Generated adjacency matrices from molecular graphs"
                )

            except Exception as e:
                logger.warning(f"Failed to generate adjacency matrices: {e}")

        if adj_matrix1 is not None and adj_matrix2 is not None:
            try:
                rmsd = self._symmrmsd(
                    pos1=pos1,
                    pos2=pos2,
                    symbols1=symbols1,
                    symbols2=symbols2,
                    adj_matrix1=adj_matrix1,
                    adj_matrix2=adj_matrix2,
                    mol_idx_pair=(i, j),  # Pass indices for mapping storage
                )
                return rmsd
            except Exception as e:
                # If symmetry-corrected RMSD fails, store no mapping and return infinity
                logger.warning(f"Symmetry-corrected RMSD failed: {e}")
                self.best_isomorphisms[(i, j)] = None
                return np.inf

        # If no adjacency matrices available, store no mapping and return infinity
        self.best_isomorphisms[(i, j)] = None
        return np.inf

    def get_best_isomorphism(
        self, mol_idx1: int, mol_idx2: int
    ) -> Optional[Tuple[List[int], List[int]]]:
        """
        Get the best isomorphism mapping between two molecules.

        Args:
            mol_idx1 (int): Index of the first molecule
            mol_idx2 (int): Index of the second molecule

        Returns:
            tuple[list, list] | None: Best isomorphism as (indices1, indices2)
                                     or None if molecules are not isomorphic

        Notes:
            This method provides access to the atom correspondence information
            computed during RMSD calculation, which is essential for understanding
            molecular similarity from a chemical perspective.
        """
        # Check both orientations as the dictionary might store (i,j) or (j,i)
        if (mol_idx1, mol_idx2) in self.best_isomorphisms:
            return self.best_isomorphisms[(mol_idx1, mol_idx2)]
        elif (mol_idx2, mol_idx1) in self.best_isomorphisms:
            # Return the reverse mapping
            mapping = self.best_isomorphisms[(mol_idx2, mol_idx1)]
            if mapping is not None:
                return mapping[1], mapping[0]  # Swap the indices
            return None
        else:
            # Mapping not computed yet
            return None

    def get_all_best_isomorphisms(self) -> dict:
        """
        Get all computed best isomorphism mappings.

        Returns:
            dict: Dictionary mapping molecule pair indices to their best isomorphisms

        Notes:
            Useful for analyzing all computed atom correspondences in the dataset.
        """
        return self.best_isomorphisms.copy()


class IRMSDGrouper(RMSDGrouper):
    """
    Invariant RMSD (iRMSD) Grouper.

    This grouper computes the permutation-invariant RMSD between molecular
    structures using canonical atom identification based on molecular graph
    topology (All-Pair-Shortest-Path algorithm, similar to Morgan algorithm).

    The implementation follows the iRMSD algorithm from:
    J. Chem. Inf. Model. 2025, 65, 4501-4511

    **Features:**
    - Canonical atom identifiers using APSP (All-Pair-Shortest-Path) algorithm
    - D3 coordination number for connectivity detection
    - Principal axes alignment using moment of inertia tensor
    - Hungarian algorithm for optimal atom permutation within equivalent groups
    - Quaternion-based RMSD calculation
    - Stereochemistry/inversion checking (for chiral molecules)
    - Full native Python implementation matching the original Fortran irmsd package

    Parameters:
        threshold (float): RMSD threshold for grouping (default: 0.125 Å)
        check_stereo (str): Control stereochemistry/inversion checking.
            - 'auto' (default): automatically detect if inversion check needed
            - 'on': force inversion check on
            - 'off': disable inversion check
        ignore_hydrogens (bool): Whether to exclude hydrogen atoms from calculation
    """

    # Standard rotation matrices for symmetry operations
    Rx180 = np.array([[1.0, 0.0, 0.0], [0.0, -1.0, 0.0], [0.0, 0.0, -1.0]])
    Ry180 = np.array([[-1.0, 0.0, 0.0], [0.0, 1.0, 0.0], [0.0, 0.0, -1.0]])
    Rz180 = np.array([[-1.0, 0.0, 0.0], [0.0, -1.0, 0.0], [0.0, 0.0, 1.0]])
    Rx90 = np.array([[1.0, 0.0, 0.0], [0.0, 0.0, 1.0], [0.0, -1.0, 0.0]])
    Ry90 = np.array([[0.0, 0.0, -1.0], [0.0, 1.0, 0.0], [1.0, 0.0, 0.0]])
    Rz90 = np.array([[0.0, 1.0, 0.0], [-1.0, 0.0, 0.0], [0.0, 0.0, 1.0]])
    Rx90T = Rx90.T
    Ry90T = Ry90.T
    Rz90T = Rz90.T

    # D3 covalent radii (in Angstroms) for connectivity detection
    # Reference: S. Grimme, J. Antony, S. Ehrlich, H. Krieg, J. Chem. Phys. 2010, 132, 154104
    _COVALENT_RADII_D3 = {
        1: 0.32,
        2: 0.46,
        3: 1.20,
        4: 0.94,
        5: 0.77,
        6: 0.75,
        7: 0.71,
        8: 0.63,
        9: 0.64,
        10: 0.67,
        11: 1.40,
        12: 1.25,
        13: 1.13,
        14: 1.04,
        15: 1.10,
        16: 1.02,
        17: 0.99,
        18: 0.96,
        19: 1.76,
        20: 1.54,
        21: 1.33,
        22: 1.22,
        23: 1.21,
        24: 1.10,
        25: 1.07,
        26: 1.04,
        27: 1.00,
        28: 0.99,
        29: 1.01,
        30: 1.09,
        31: 1.12,
        32: 1.09,
        33: 1.15,
        34: 1.10,
        35: 1.14,
        36: 1.17,
        37: 1.89,
        38: 1.67,
        39: 1.47,
        40: 1.39,
        41: 1.32,
        42: 1.24,
        43: 1.15,
        44: 1.13,
        45: 1.13,
        46: 1.08,
        47: 1.15,
        48: 1.23,
        49: 1.28,
        50: 1.26,
        51: 1.26,
        52: 1.23,
        53: 1.32,
        54: 1.31,
        55: 2.09,
        56: 1.76,
        57: 1.62,
        58: 1.47,
        59: 1.58,
        60: 1.57,
        61: 1.56,
        62: 1.55,
        63: 1.51,
        64: 1.52,
        65: 1.51,
        66: 1.50,
        67: 1.49,
        68: 1.49,
        69: 1.48,
        70: 1.53,
        71: 1.46,
        72: 1.37,
        73: 1.31,
        74: 1.23,
        75: 1.18,
        76: 1.16,
        77: 1.11,
        78: 1.12,
        79: 1.13,
        80: 1.32,
        81: 1.30,
        82: 1.30,
        83: 1.36,
        84: 1.31,
        85: 1.38,
        86: 1.42,
        87: 2.01,
        88: 1.81,
        89: 1.67,
        90: 1.58,
        91: 1.52,
        92: 1.53,
        93: 1.54,
        94: 1.55,
    }

    # D3 parameters for coordination number calculation (k1, k2 in the CN formula)
    _CN_K1 = 16.0
    _CN_K2 = 4.0 / 3.0

    def __init__(
        self,
        molecules: Iterable[Molecule],
        threshold=None,
        num_groups=None,
        num_procs: int = 1,
        align_molecules: bool = False,
        ignore_hydrogens: bool = False,
        check_stereo: str = "auto",
        label: str = None,
        **kwargs,
    ):
        """
        Initialize iRMSD grouper.

        Args:
            molecules: Collection of molecules to group
            threshold: RMSD threshold for grouping (default: 0.125 Angstroms)
            num_groups (int): Number of groups to create (alternative to threshold)
            num_procs: Number of processes for parallel computation
            align_molecules: Whether to align molecules (legacy parameter, ignored)
            ignore_hydrogens: Whether to exclude hydrogen atoms
            check_stereo (str): Control stereochemistry/inversion checking.
                - 'auto' (default): automatically detect if inversion check needed
                - 'on': force inversion check on
                - 'off': disable inversion check
            label (str): Label for output files
        """
        # Set default threshold if not provided
        if threshold is None and num_groups is None:
            threshold = 0.125  # Default iRMSD threshold in Angstroms

        super().__init__(
            molecules,
            threshold,
            num_groups,
            num_procs,
            align_molecules,
            ignore_hydrogens,
            label=label,
        )
        self._pt = PeriodicTable()

        # Set _check_stereo_enabled based on check_stereo flag
        # check_stereo: 'auto', 'on', 'off'
        self.check_stereo = (
            check_stereo.lower() if isinstance(check_stereo, str) else "auto"
        )
        if self.check_stereo == "off":
            self._check_stereo_enabled = False
        else:
            # For 'auto' and 'on', enable stereo check
            self._check_stereo_enabled = True

        logger.info(
            "Using native Python iRMSD implementation with APSP algorithm"
        )

    # ========== D3 Coordination Number Calculation ==========

    def _get_covalent_radius(self, atomic_number: int) -> float:
        """Get D3 covalent radius for an atomic number."""
        return self._COVALENT_RADII_D3.get(
            atomic_number, 1.5
        )  # Default fallback

    def _compute_cn_d3(
        self, atomic_numbers: np.ndarray, positions: np.ndarray
    ) -> np.ndarray:
        """
        Compute D3 coordination numbers for all atoms.

        This implements the D3 CN formula:
        CN_i = sum_j [1 / (1 + exp(-k1 * (k2 * (R_cov_i + R_cov_j) / r_ij - 1)))]

        Args:
            atomic_numbers: (N,) array of atomic numbers
            positions: (N, 3) array of atomic positions

        Returns:
            cn: (N,) array of coordination numbers
        """
        n_atoms = len(atomic_numbers)
        cn = np.zeros(n_atoms, dtype=np.float64)

        # Get covalent radii for all atoms
        cov_radii = np.array(
            [self._get_covalent_radius(z) for z in atomic_numbers]
        )

        # Compute pairwise distances
        for i in range(n_atoms):
            for j in range(i + 1, n_atoms):
                rij = np.linalg.norm(positions[i] - positions[j])
                if rij < 1e-6:
                    continue

                # Reference covalent distance
                rcov_sum = cov_radii[i] + cov_radii[j]

                # D3 CN counting function
                arg = -self._CN_K1 * (self._CN_K2 * rcov_sum / rij - 1.0)
                cn_contrib = 1.0 / (1.0 + np.exp(arg))

                cn[i] += cn_contrib
                cn[j] += cn_contrib

        return cn

    def _build_connectivity_matrix(
        self,
        atomic_numbers: np.ndarray,
        positions: np.ndarray,
        cn_threshold: float = 0.5,
    ) -> np.ndarray:
        """
        Build connectivity (adjacency) matrix from atomic positions using D3-like bond detection.

        Two atoms are considered bonded if their contribution to CN exceeds cn_threshold.

        Args:
            atomic_numbers: (N,) array of atomic numbers
            positions: (N, 3) array of atomic positions
            cn_threshold: threshold for bond detection (default 0.5)

        Returns:
            connectivity: (N, N) integer matrix (1 = bonded, 0 = not bonded)
        """
        n_atoms = len(atomic_numbers)
        connectivity = np.zeros((n_atoms, n_atoms), dtype=np.int32)

        # Get covalent radii for all atoms
        cov_radii = np.array(
            [self._get_covalent_radius(z) for z in atomic_numbers]
        )

        for i in range(n_atoms):
            for j in range(i + 1, n_atoms):
                rij = np.linalg.norm(positions[i] - positions[j])
                if rij < 1e-6:
                    continue

                rcov_sum = cov_radii[i] + cov_radii[j]

                # D3 CN counting function
                arg = -self._CN_K1 * (self._CN_K2 * rcov_sum / rij - 1.0)
                cn_contrib = 1.0 / (1.0 + np.exp(arg))

                if cn_contrib > cn_threshold:
                    connectivity[i, j] = 1
                    connectivity[j, i] = 1

        return connectivity

    # ========== APSP Algorithm for Canonical Ranking ==========

    def _floyd_warshall_apsp(self, connectivity: np.ndarray) -> np.ndarray:
        """
        Compute All-Pairs Shortest Paths using Floyd-Warshall algorithm.

        Args:
            connectivity: (N, N) adjacency matrix (1 = connected, 0 = not connected)

        Returns:
            dist: (N, N) shortest path distance matrix
        """
        n = connectivity.shape[0]
        INF = n + 1  # Infinity for unreachable pairs

        # Initialize distance matrix
        dist = np.full((n, n), INF, dtype=np.int32)
        np.fill_diagonal(dist, 0)

        # Set direct connections to distance 1
        dist[connectivity == 1] = 1

        # Floyd-Warshall algorithm
        for k in range(n):
            for i in range(n):
                for j in range(n):
                    if dist[i, k] + dist[k, j] < dist[i, j]:
                        dist[i, j] = dist[i, k] + dist[k, j]

        return dist

    def _compute_apsp_invariants(
        self, atomic_numbers: np.ndarray, dist_matrix: np.ndarray
    ) -> np.ndarray:
        """
        Compute APSP+ invariants for each atom.

        The APSP+ invariant for atom i is a tuple of:
        - atomic number
        - sorted list of (distance, neighbor_atomic_number) for all atoms

        This is simplified to a hash for comparison purposes.

        Args:
            atomic_numbers: (N,) array of atomic numbers
            dist_matrix: (N, N) shortest path distance matrix

        Returns:
            invariants: (N,) array of invariant values (hashes)
        """
        n = len(atomic_numbers)
        invariants = np.zeros(n, dtype=np.int64)

        # Use Python integers to avoid numpy overflow
        P = 31  # Prime base
        M = 10**18 + 9  # Large prime modulus

        for i in range(n):
            # Build invariant tuple: (atomic_number, sorted list of (dist, neighbor_Z))
            inv_list = []
            for j in range(n):
                if i != j:
                    inv_list.append(
                        (int(dist_matrix[i, j]), int(atomic_numbers[j]))
                    )

            # Sort by distance first, then by atomic number
            inv_list.sort()

            # Create a hash from the invariant
            # Use polynomial hashing: hash = sum(a[i] * p^i) mod M
            # where a[i] combines distance and atomic number
            h = int(
                atomic_numbers[i]
            )  # Start with own atomic number (Python int)
            p_pow = P
            for d, z in inv_list:
                # Use Python integers to avoid overflow
                h = (h + (d * 1000 + z) * p_pow) % M
                p_pow = (p_pow * P) % M

            invariants[i] = h

        return invariants

    def _compute_canonical_ranks(
        self, atomic_numbers: np.ndarray, positions: np.ndarray
    ) -> np.ndarray:
        """
        Compute canonical ranks for atoms using APSP algorithm.

        Atoms with the same rank are topologically equivalent and can be permuted.

        Args:
            atomic_numbers: (N,) array of atomic numbers
            positions: (N, 3) array of atomic positions

        Returns:
            ranks: (N,) array of canonical ranks (1-indexed, same rank = equivalent)
        """

        # Build connectivity matrix
        connectivity = self._build_connectivity_matrix(
            atomic_numbers, positions
        )

        # Compute APSP distance matrix
        dist_matrix = self._floyd_warshall_apsp(connectivity)

        # Compute APSP+ invariants
        invariants = self._compute_apsp_invariants(atomic_numbers, dist_matrix)

        # Assign ranks based on unique invariants
        unique_invs = sorted(set(invariants))
        inv_to_rank = {inv: rank + 1 for rank, inv in enumerate(unique_invs)}

        ranks = np.array(
            [inv_to_rank[inv] for inv in invariants], dtype=np.int32
        )

        return ranks

    # ========== Core iRMSD Utilities ==========

    def _get_atomic_masses(self, symbols):
        """Get atomic masses for given element symbols."""
        return np.array([self._pt.to_atomic_mass(s) for s in symbols])

    def _compute_cma_and_shift(self, coords, masses):
        """Compute mass-weighted center of mass and shift coordinates."""
        total_mass = np.sum(masses) + 1e-20
        cma = np.sum(coords * masses[:, np.newaxis], axis=0) / total_mass
        return coords - cma, cma

    def _compute_inertia_tensor(self, coords, masses):
        """Compute the moment of inertia tensor."""
        t = np.array([i * 1e-10 for i in range(1, 7)])
        for m, r in zip(masses, coords):
            x, y, z = r
            t[0] += m * (y**2 + z**2)
            t[1] -= m * x * y
            t[2] += m * (z**2 + x**2)
            t[3] -= m * z * x
            t[4] -= m * y * z
            t[5] += m * (x**2 + y**2)
        return np.array(
            [[t[0], t[1], t[3]], [t[1], t[2], t[4]], [t[3], t[4], t[5]]]
        )

    def _compute_rotational_constants(self, inertia):
        """Compute rotational constants from inertia tensor."""
        icm2MHz = 2.9979245e4
        Aamu2icm = 16.8576522
        eig, evec = np.linalg.eigh(inertia)
        evec[np.abs(evec) < 1e-9] = 0.0
        rot = np.zeros(3)
        for i in range(3):
            if eig[i] < 3e-4:
                rot[i] = 0.0
            else:
                rot[i] = icm2MHz * Aamu2icm / eig[i]
        return rot, evec

    def _check_unique_axes(self, rot, thr=0.01):
        """Check uniqueness of rotational constants."""
        unique = np.array([False, False, False])
        if rot[0] < 1e-10 or rot[1] < 1e-10:
            return unique, 3
        diff = np.array(
            [
                abs(rot[1] / rot[0] - 1.0),
                abs(rot[2] / rot[0] - 1.0),
                abs(rot[2] / rot[1] - 1.0),
            ]
        )
        if diff[0] > thr and diff[1] > thr:
            unique[0] = True
        if diff[0] > thr and diff[2] > thr:
            unique[1] = True
        if diff[1] > thr and diff[2] > thr:
            unique[2] = True
        n_unique = np.sum(unique)
        if n_unique == 3:
            return unique, 0
        elif n_unique == 1:
            if unique[0]:
                return unique, 1
            elif unique[2]:
                return unique, 2
        return unique, 3

    def _align_to_principal_axes(self, coords, masses):
        """Align molecule to principal axes."""
        shifted_coords, _ = self._compute_cma_and_shift(coords, masses)
        inertia = self._compute_inertia_tensor(shifted_coords, masses)
        rot, evec = self._compute_rotational_constants(inertia)
        if np.linalg.det(evec) < 0:
            evec[:, 0] = -evec[:, 0]
        return shifted_coords @ evec, rot, evec

    def _element_permutation(self, P, Q, symbols):
        """Find optimal permutation within element groups using Hungarian algorithm."""
        n_atoms = len(symbols)
        perm = np.arange(n_atoms)
        total_cost = 0.0
        element_groups = defaultdict(list)
        for i, symbol in enumerate(symbols):
            element_groups[symbol].append(i)
        for element, indices in element_groups.items():
            if len(indices) <= 1:
                continue
            indices = np.array(indices)
            cost_matrix = cdist(P[indices], Q[indices], metric="sqeuclidean")
            row_ind, col_ind = linear_sum_assignment(cost_matrix)
            perm[indices[row_ind]] = indices[col_ind]
            total_cost += cost_matrix[row_ind, col_ind].sum()
        return perm, total_cost

    def _rmsd_quaternion(self, ref_coords, mol_coords):
        """Calculate RMSD using quaternion-based algorithm."""
        n_atoms = len(ref_coords)
        x = ref_coords.T.copy()
        y = mol_coords.T.copy()
        x = x - np.mean(x, axis=1, keepdims=True)
        y = y - np.mean(y, axis=1, keepdims=True)
        x_norm = np.sum(x**2)
        y_norm = np.sum(y**2)
        R = x @ y.T
        S = np.array(
            [
                [
                    R[0, 0] + R[1, 1] + R[2, 2],
                    R[1, 2] - R[2, 1],
                    R[2, 0] - R[0, 2],
                    R[0, 1] - R[1, 0],
                ],
                [
                    R[1, 2] - R[2, 1],
                    R[0, 0] - R[1, 1] - R[2, 2],
                    R[0, 1] + R[1, 0],
                    R[0, 2] + R[2, 0],
                ],
                [
                    R[2, 0] - R[0, 2],
                    R[0, 1] + R[1, 0],
                    -R[0, 0] + R[1, 1] - R[2, 2],
                    R[1, 2] + R[2, 1],
                ],
                [
                    R[0, 1] - R[1, 0],
                    R[0, 2] + R[2, 0],
                    R[1, 2] + R[2, 1],
                    -R[0, 0] - R[1, 1] + R[2, 2],
                ],
            ]
        )
        eigenvalues, _ = np.linalg.eigh(S)
        lambda_max = eigenvalues[-1]
        rmsd_sq = max(0.0, (x_norm + y_norm - 2.0 * lambda_max)) / n_atoms
        return np.sqrt(rmsd_sq)

    def _test_orientations(
        self, ref_coords, mol_coords, symbols, uniqueness_case
    ):
        """Test multiple orientations and return the best RMSD."""
        best_rmsd = np.inf
        best_perm = None
        n_iterations = {0: 1, 1: 2, 2: 2}.get(uniqueness_case, 4)
        mol_working = mol_coords.copy()

        for ii in range(n_iterations):
            for rot_matrix in [None, self.Rx180, self.Ry180, self.Rx180]:
                if rot_matrix is not None:
                    mol_working = mol_working @ rot_matrix.T
                perm, _ = self._element_permutation(
                    ref_coords, mol_working, symbols
                )
                rmsd = self._rmsd_quaternion(ref_coords, mol_working[perm])
                if rmsd < best_rmsd:
                    best_rmsd = rmsd
                    best_perm = perm.copy()
                if best_rmsd < 1e-10:
                    return best_rmsd, best_perm
            mol_working = mol_working @ self.Ry180.T
            if uniqueness_case == 0:
                break
            elif uniqueness_case == 1 and ii == 0:
                mol_working = mol_working @ self.Rx90.T
            elif uniqueness_case == 2 and ii == 0:
                mol_working = mol_working @ self.Rz90.T
            elif uniqueness_case == 3:
                if ii == 0:
                    mol_working = mol_working @ self.Rz90.T
                elif ii == 1:
                    mol_working = mol_working @ self.Rz90T.T @ self.Ry90.T
                elif ii == 2:
                    mol_working = mol_working @ self.Ry90T.T @ self.Rx90.T
        return best_rmsd, best_perm

    # ========== APSP-based Permutation ==========

    def _apsp_permutation(self, ref_coords, mol_coords, atomic_numbers, ranks):
        """
        Find optimal permutation using APSP-based canonical ranking.

        Within each equivalence class (atoms with same rank), use Hungarian algorithm
        to find the optimal atom-to-atom assignment.

        Args:
            ref_coords: (N, 3) reference coordinates
            mol_coords: (N, 3) molecule coordinates
            atomic_numbers: (N,) atomic numbers
            ranks: (N,) canonical ranks from APSP

        Returns:
            perm: (N,) optimal permutation array
            cost: total assignment cost
        """
        n_atoms = len(atomic_numbers)
        perm = np.arange(n_atoms)
        total_cost = 0.0

        # Group atoms by their canonical rank
        rank_groups = defaultdict(list)
        for i, rank in enumerate(ranks):
            rank_groups[rank].append(i)

        # For each equivalence class, use Hungarian algorithm
        for rank, indices in rank_groups.items():
            if len(indices) <= 1:
                continue

            indices = np.array(indices)
            cost_matrix = cdist(
                ref_coords[indices], mol_coords[indices], metric="sqeuclidean"
            )
            row_ind, col_ind = linear_sum_assignment(cost_matrix)
            perm[indices[row_ind]] = indices[col_ind]
            total_cost += cost_matrix[row_ind, col_ind].sum()

        return perm, total_cost

    def _test_orientations_apsp(
        self, ref_coords, mol_coords, atomic_numbers, ranks, uniqueness_case
    ):
        """
        Test multiple orientations using APSP-based permutation and return best RMSD.

        This is the core iRMSD algorithm that:
        1. Tests multiple molecular orientations (rotations around principal axes)
        2. Uses APSP canonical ranking to identify equivalent atoms
        3. Uses Hungarian algorithm within equivalence classes for optimal matching
        4. Computes RMSD using quaternion algorithm

        Args:
            ref_coords: reference molecule coordinates (already aligned to principal axes)
            mol_coords: test molecule coordinates (already aligned to principal axes)
            atomic_numbers: (N,) atomic numbers
            ranks: (N,) canonical ranks from APSP algorithm
            uniqueness_case: symmetry case (0=asymmetric, 1-3=various symmetries)

        Returns:
            best_rmsd: best RMSD found across all orientations
            best_perm: best permutation found
        """
        best_rmsd = np.inf
        best_perm = None

        # Number of major orientation iterations depends on symmetry
        n_iterations = {0: 1, 1: 2, 2: 2}.get(uniqueness_case, 4)
        mol_working = mol_coords.copy()

        for ii in range(n_iterations):
            # Test 4 rotations: identity, Rx180, Ry180, Rx180 (equivalent to Rz180)
            for rot_matrix in [None, self.Rx180, self.Ry180, self.Rx180]:
                if rot_matrix is not None:
                    mol_working = mol_working @ rot_matrix.T

                # Use APSP-based permutation
                perm, _ = self._apsp_permutation(
                    ref_coords, mol_working, atomic_numbers, ranks
                )

                # Calculate RMSD with optimal permutation
                rmsd = self._rmsd_quaternion(ref_coords, mol_working[perm])

                if rmsd < best_rmsd:
                    best_rmsd = rmsd
                    best_perm = perm.copy()

                # Early termination if we found perfect match
                if best_rmsd < 1e-10:
                    return best_rmsd, best_perm

            # Apply additional rotation for symmetric cases
            mol_working = mol_working @ self.Ry180.T

            if uniqueness_case == 0:
                break
            elif uniqueness_case == 1 and ii == 0:
                mol_working = mol_working @ self.Rx90.T
            elif uniqueness_case == 2 and ii == 0:
                mol_working = mol_working @ self.Rz90.T
            elif uniqueness_case == 3:
                if ii == 0:
                    mol_working = mol_working @ self.Rz90.T
                elif ii == 1:
                    mol_working = mol_working @ self.Rz90T.T @ self.Ry90.T
                elif ii == 2:
                    mol_working = mol_working @ self.Ry90T.T @ self.Rx90.T

        return best_rmsd, best_perm

    def _irmsd_core(self, pos1, pos2, atomic_numbers1, atomic_numbers2):
        """
        Core iRMSD calculation using APSP algorithm.

        This is the main algorithm that matches the original Fortran irmsd package:
        1. Compute canonical ranks using APSP (All-Pairs Shortest Paths) for both molecules
        2. Find correspondence between atoms based on canonical ranks
        3. Align both molecules to their principal axes
        4. Test multiple orientations with APSP-based atom permutation
        5. Optionally check inverted structure if stereo_check is enabled

        Args:
            pos1: (N, 3) reference molecule positions
            pos2: (N, 3) test molecule positions
            atomic_numbers1: (N,) atomic numbers for reference molecule
            atomic_numbers2: (N,) atomic numbers for test molecule

        Returns:
            float: best iRMSD value
        """

        # Get atomic masses for principal axis alignment
        # Note: to_atomic_mass expects element symbols, but we have atomic numbers
        # Convert atomic numbers to symbols first using to_symbol method
        symbols1 = [self._pt.to_symbol(int(z)) for z in atomic_numbers1]
        symbols2 = [self._pt.to_symbol(int(z)) for z in atomic_numbers2]
        masses1 = np.array([self._pt.to_atomic_mass(s) for s in symbols1])
        masses2 = np.array([self._pt.to_atomic_mass(s) for s in symbols2])

        # Compute canonical ranks using APSP algorithm for both molecules
        ranks1 = self._compute_canonical_ranks(atomic_numbers1, pos1)
        ranks2 = self._compute_canonical_ranks(atomic_numbers2, pos2)

        # Find initial atom correspondence based on canonical ranks
        # This maps atoms from mol2 to mol1 based on their canonical signatures
        initial_perm = self._find_initial_correspondence(
            atomic_numbers1, ranks1, atomic_numbers2, ranks2
        )

        if initial_perm is None:
            # Cannot establish correspondence - molecules may have different topology
            return np.inf

        # Reorder mol2 according to initial correspondence
        pos2_reordered = pos2[initial_perm]
        masses2_reordered = masses2[initial_perm]

        # Now both molecules should have corresponding atoms in the same order
        # Use ranks from molecule 1 for the Hungarian algorithm
        ranks = ranks1

        # Align both molecules to principal axes
        ref_aligned, rot_ref, evec_ref = self._align_to_principal_axes(
            pos1, masses1
        )
        mol_aligned, rot_mol, evec_mol = self._align_to_principal_axes(
            pos2_reordered, masses2_reordered
        )

        # Check uniqueness of principal axes
        _, uniqueness_case = self._check_unique_axes(rot_ref)

        # Test orientations with APSP-based permutation
        best_rmsd, _ = self._test_orientations_apsp(
            ref_aligned, mol_aligned, atomic_numbers1, ranks, uniqueness_case
        )

        # Check inverted structure if check_stereo is enabled
        if self._check_stereo_enabled:
            # Invert the second molecule
            mol_inverted = -mol_aligned

            # Test orientations with inverted structure
            rmsd_inv, _ = self._test_orientations_apsp(
                ref_aligned,
                mol_inverted,
                atomic_numbers1,
                ranks,
                uniqueness_case,
            )

            if rmsd_inv < best_rmsd:
                best_rmsd = rmsd_inv

        return best_rmsd

    def _find_initial_correspondence(
        self, atomic_numbers1, ranks1, atomic_numbers2, ranks2
    ):
        """
        Find initial atom correspondence between two molecules based on canonical ranks.

        For each canonical rank group, we need to match atoms from mol2 to mol1.
        If both molecules have the same number of atoms with the same (rank, atomic_number)
        signatures, we can establish a correspondence.

        Args:
            atomic_numbers1: (N,) atomic numbers for molecule 1
            ranks1: (N,) canonical ranks for molecule 1
            atomic_numbers2: (N,) atomic numbers for molecule 2
            ranks2: (N,) canonical ranks for molecule 2

        Returns:
            perm: (N,) permutation array mapping mol2 indices to mol1 positions,
                  or None if correspondence cannot be established
        """
        n_atoms = len(atomic_numbers1)
        perm = np.zeros(n_atoms, dtype=np.int32)

        # Create signature tuples (rank, atomic_number) for each atom
        sig1 = [(ranks1[i], atomic_numbers1[i]) for i in range(n_atoms)]
        sig2 = [(ranks2[i], atomic_numbers2[i]) for i in range(n_atoms)]

        # Group atoms by signature for mol1
        sig_to_indices1 = defaultdict(list)
        for i, sig in enumerate(sig1):
            sig_to_indices1[sig].append(i)

        # Group atoms by signature for mol2
        sig_to_indices2 = defaultdict(list)
        for i, sig in enumerate(sig2):
            sig_to_indices2[sig].append(i)

        # Check that both molecules have the same signature distribution
        if set(sig_to_indices1.keys()) != set(sig_to_indices2.keys()):
            return None

        for sig in sig_to_indices1:
            if len(sig_to_indices1[sig]) != len(sig_to_indices2[sig]):
                return None

        # Assign atoms from mol2 to mol1 positions
        # For atoms with the same signature, assign them in order
        # (the Hungarian algorithm will later optimize within groups)
        for sig in sig_to_indices1:
            indices1 = sig_to_indices1[sig]
            indices2 = sig_to_indices2[sig]

            for idx1, idx2 in zip(indices1, indices2):
                perm[idx1] = idx2

        return perm

    def _calculate_rmsd(self, mol_idx_pair: Tuple[int, int]) -> float:
        """
        Calculate iRMSD between two molecules.

        This method overrides the base class _calculate_rmsd to use the
        iRMSD algorithm with APSP-based canonical ranking.

        Args:
            mol_idx_pair: Tuple of (index1, index2) for the two molecules

        Returns:
            float: iRMSD value, or np.inf if calculation fails
        """
        i, j = mol_idx_pair
        mol1 = self.molecules[i]
        mol2 = self.molecules[j]

        # Get positions and symbols
        pos1 = mol1.positions
        pos2 = mol2.positions
        symbols1 = mol1.chemical_symbols
        symbols2 = mol2.chemical_symbols

        # Validate that molecules have same atoms (sorted comparison to allow different ordering)
        if sorted(symbols1) != sorted(symbols2):
            logger.warning(f"Molecules {i} and {j} have different atom types")
            return np.inf

        if pos1.shape != pos2.shape:
            logger.warning(f"Molecules {i} and {j} have different shapes")
            return np.inf

        # Filter hydrogens if requested
        if self.ignore_hydrogens:
            non_h_mask1 = np.array([s != "H" for s in symbols1])
            non_h_mask2 = np.array([s != "H" for s in symbols2])
            pos1 = pos1[non_h_mask1]
            pos2 = pos2[non_h_mask2]
            symbols1 = [s for s in symbols1 if s != "H"]
            symbols2 = [s for s in symbols2 if s != "H"]

        # Convert symbols to atomic numbers for both molecules
        atomic_numbers1 = np.array(
            [self._pt.to_atomic_number(s) for s in symbols1]
        )
        atomic_numbers2 = np.array(
            [self._pt.to_atomic_number(s) for s in symbols2]
        )

        # Calculate iRMSD using APSP core algorithm
        return self._irmsd_core(pos1, pos2, atomic_numbers1, atomic_numbers2)


class PymolRMSDGrouper(RMSDGrouper):
    """
    Group molecules using PyMOL's align command for RMSD calculation.

    This grouper writes molecules to temporary XYZ files and uses PyMOL's
    align command to calculate RMSD values, following the same approach
    as PyMOLAlignJobRunner in jobs/mol/runner.py.

    Simple workflow:
    1. Write molecules to temporary XYZ files
    2. Load into PyMOL using cmd.load()
    3. Run align command: align mol_i, mol_j (with PyMOL default parameters)
    4. Get RMSD from the result

    PyMOL align default parameters (with outlier rejection):
    - cutoff=2.0: outlier rejection cutoff in RMS
    - cycles=5: maximum number of outlier rejection cycles
    """

    def __init__(
        self,
        molecules: Iterable[Molecule],
        threshold: float = 0.5,
        num_groups=None,
        num_procs: int = 1,
        ignore_hydrogens: bool = False,
        label: str = None,
        **kwargs,
    ):
        """
        Initialize PyMOL RMSD grouper.

        Args:
            molecules: Collection of molecules to group
            threshold: RMSD threshold for grouping (Angstroms)
            num_groups: Number of groups to create (alternative to threshold)
            num_procs: Number of processes (PyMOL uses single process)
            ignore_hydrogens: Whether to exclude hydrogen atoms
            label: Label for output files

        Note:
            Uses PyMOL's default align parameters including outlier rejection
            (cutoff=2.0, cycles=5). This matches PyMOL's standard behavior.
        """
        # Note: align_molecules is always True for PyMOL align
        super().__init__(
            molecules=molecules,
            threshold=threshold,
            num_groups=num_groups,
            num_procs=1,  # PyMOL is single-threaded
            align_molecules=True,
            ignore_hydrogens=ignore_hydrogens,
            label=label,
            **kwargs,
        )
        self._temp_dir = None
        self._xyz_files = []
        self._mol_names = []
        self._alignment_cache = {}

        # Initialize PyMOL and prepare molecules
        self._init_pymol()
        self._prepare_molecules()

    def _init_pymol(self):
        """Initialize PyMOL in command-line mode."""
        try:
            import pymol
            from pymol import cmd

            # Initialize PyMOL in quiet command-line mode
            pymol.finish_launching(["pymol", "-qc"])
            self.cmd = cmd
            self.cmd.reinitialize()
        except ImportError as e:
            raise ImportError(
                f"PyMOL not available: {e}\n"
                f"Please install PyMOL: conda install -c conda-forge pymol-open-source"
            )

    def _prepare_molecules(self):
        """Write all molecules to temporary XYZ files and load into PyMOL."""
        import os
        import tempfile

        # Create temporary directory
        self._temp_dir = tempfile.mkdtemp(prefix="pymol_rmsd_")

        # Write each molecule to XYZ file and load into PyMOL
        for i, mol in enumerate(self.molecules):
            mol_name = f"mol_{i}"
            xyz_path = os.path.join(self._temp_dir, f"{mol_name}.xyz")

            # Write XYZ file
            self._write_xyz(mol, xyz_path)

            # Load into PyMOL
            self.cmd.load(xyz_path, mol_name)

            self._xyz_files.append(xyz_path)
            self._mol_names.append(mol_name)

    def _write_xyz(self, mol: Molecule, filepath: str):
        """Write molecule to XYZ format file."""
        positions = mol.positions
        symbols = mol.chemical_symbols

        if self.ignore_hydrogens:
            # Filter out hydrogens
            non_h_indices = [i for i, s in enumerate(symbols) if s != "H"]
            positions = positions[non_h_indices]
            symbols = [symbols[i] for i in non_h_indices]

        with open(filepath, "w") as f:
            f.write(f"{len(symbols)}\n")
            f.write("Generated by PymolRMSDGrouper\n")
            for pos, sym in zip(positions, symbols):
                f.write(f"{sym} {pos[0]:.6f} {pos[1]:.6f} {pos[2]:.6f}\n")

    def _calculate_rmsd(self, idx_pair: Tuple[int, int]) -> float:
        """
        Calculate RMSD between two molecules using PyMOL align.

        Simply runs: align mol_i, mol_j
        with PyMOL default parameters (cutoff=2.0, cycles=5 for outlier rejection).
        Returns the RMSD value after alignment and outlier rejection.
        """
        i, j = idx_pair

        # Check cache
        cache_key = tuple(sorted([i, j]))
        if cache_key in self._alignment_cache:
            return self._alignment_cache[cache_key]

        mol_i = self._mol_names[i]
        mol_j = self._mol_names[j]

        try:
            # Run PyMOL align command with default parameters
            # PyMOL defaults: cutoff=2.0, cycles=5 (outlier rejection enabled)
            result = self.cmd.align(mol_i, mol_j)

            if result is None or not isinstance(result, (list, tuple)):
                rmsd = float("inf")
            else:
                rmsd = result[0]  # First element is RMSD after alignment

                # Validate RMSD
                if rmsd < 0 or not np.isfinite(rmsd):
                    rmsd = float("inf")

        except Exception as e:
            logger.warning(f"PyMOL align failed for ({i}, {j}): {e}")
            rmsd = float("inf")

        self._alignment_cache[cache_key] = rmsd
        return rmsd

    def __del__(self):
        """Clean up temporary files and PyMOL session."""
        import shutil

        # Clean up PyMOL
        try:
            if hasattr(self, "cmd"):
                self.cmd.quit()
        except Exception:
            pass

        # Clean up temp directory
        try:
            if hasattr(self, "_temp_dir") and self._temp_dir:
                shutil.rmtree(self._temp_dir, ignore_errors=True)
        except Exception:
            pass

    def __repr__(self):
        if self.num_groups is not None:
            return f"{self.__class__.__name__}(num_groups={self.num_groups})"
        else:
            return f"{self.__class__.__name__}(threshold={self.threshold})"


class RMSDGrouperSharedMemory(MoleculeGrouper):
    """
    Group molecules based on RMSD using shared memory optimization.

    Optimized version of RMSDGrouper that uses shared memory to reduce
    data copying overhead in multiprocessing scenarios. Provides faster
    computation for large datasets by minimizing memory allocation and
    inter-process communication costs.

    Attributes:
        molecules (Iterable[Molecule]): Inherited; collection of molecules to
            group.
        num_procs (int): Inherited; number of worker processes.
        threshold (float): RMSD threshold for grouping molecules.
        align_molecules (bool): Whether to align molecules before RMSD
            calculation.
        ignore_hydrogens (bool): Whether to exclude hydrogen atoms from RMSD.
    """

    def __init__(
        self,
        molecules: Iterable[Molecule],
        threshold: float = 0.5,  # RMSD threshold for grouping
        num_procs: int = 1,
        align_molecules: bool = True,
        ignore_hydrogens: bool = False,
        label: str = None,
        **kwargs,
    ):
        """
        Initialize RMSD grouper with shared memory optimization.

        Args:
            molecules (Iterable[Molecule]): Collection of molecules to group.
            threshold (float): RMSD threshold for grouping. Defaults to 0.5.
            num_procs (int): Number of processes for parallel computation.
            align_molecules (bool): Whether to align molecules using Kabsch
                algorithm before RMSD calculation. Defaults to True.
            ignore_hydrogens (bool): Whether to exclude hydrogen atoms from
                RMSD calculation. Defaults to False.
            label (str): Label/name for output files.
        """
        super().__init__(molecules, num_procs, label=label)
        self.threshold = threshold
        self.align_molecules = align_molecules
        self.ignore_hydrogens = ignore_hydrogens

    def group(self) -> Tuple[List[List[Molecule]], List[List[int]]]:
        """
        Group molecules using shared memory with optimized parallelism.

        Uses RawArray shared memory to minimize data copying between processes.
        Molecular positions are stored once in shared memory and accessed
        by worker processes for RMSD calculations.

        Returns:
            Tuple[List[List[Molecule]], List[List[int]]]: Tuple containing:
                - List of molecule groups (each group is a list of molecules)
                - List of index groups (corresponding indices for each group)
        """
        n = len(self.molecules)
        num_atoms = self.molecules[0].positions.shape[0]

        # 🧠 **1️⃣ Create Shared Memory (RawArray - Faster, Less Locking)**
        shared_pos = RawArray("d", n * num_atoms * 3)  # 'd' -> float64

        # Convert RawArray into numpy view
        pos_np = np.frombuffer(shared_pos, dtype=np.float64).reshape(
            n, num_atoms, 3
        )

        # Copy molecular positions into shared memory (only once!)
        for i, mol in enumerate(self.molecules):
            pos_np[i] = mol.positions

        # 🏃‍♂️ **2️⃣ Run Parallel RMSD Calculation Using Explicit Shared Memory**
        indices = [(i, j) for i in range(n) for j in range(i + 1, n)]
        with multiprocessing.Pool(
            self.num_procs,
            initializer=self._init_worker,
            initargs=(shared_pos, (n, num_atoms, 3)),
        ) as pool:
            rmsd_values = pool.map(self._calculate_rmsd, indices)

        # 🏗️ **3️⃣ Construct Adjacency Matrix for Clustering**
        adj_matrix = np.zeros((n, n), dtype=bool)
        for (i, j), rmsd in zip(indices, rmsd_values):
            if rmsd < self.threshold:
                adj_matrix[i, j] = adj_matrix[j, i] = True

        # 🔗 **4️⃣ Complete Linkage Grouping**
        # A structure joins a group only if its RMSD to ALL existing members is below the threshold
        groups, index_groups = self._complete_linkage_grouping(adj_matrix, n)

        # Cache the results to avoid recomputation
        self._cached_groups = groups
        self._cached_group_indices = index_groups

        return groups, index_groups

    def _complete_linkage_grouping(self, adj_matrix, n):
        """
        Perform complete linkage grouping (from last to first).

        A structure joins a group only if its RMSD to ALL existing members
        of that group is below the threshold (i.e., all edges exist in adj_matrix).
        This prevents the chaining effect where dissimilar structures end up
        in the same group through intermediate "bridge" structures.

        Iterates from last to first structure so that higher-energy structures
        (typically at the end) are more likely to form singleton groups,
        preserving lower-energy structures in larger groups.

        Args:
            adj_matrix: Boolean adjacency matrix where adj_matrix[i,j] = True
                       if RMSD between i and j is below threshold
            n: Number of molecules

        Returns:
            Tuple of (groups, index_groups)
        """
        assigned = [False] * n
        groups = []
        index_groups = []

        # Iterate from last to first (higher energy structures first as seeds)
        for i in range(n - 1, -1, -1):
            if assigned[i]:
                continue

            # Start a new group with molecule i
            current_group = [i]
            assigned[i] = True

            # Try to add unassigned molecules with lower indices (lower energy)
            for j in range(i - 1, -1, -1):
                if assigned[j]:
                    continue

                # Check if j is connected to ALL members in current_group
                can_join = all(
                    adj_matrix[j, member] for member in current_group
                )

                if can_join:
                    current_group.append(j)
                    assigned[j] = True

            # Sort indices within group (lowest index = lowest energy first)
            current_group.sort()
            # Add the completed group
            groups.append([self.molecules[idx] for idx in current_group])
            index_groups.append(current_group)

        # Reverse groups so that groups containing lower-energy structures come first
        groups.reverse()
        index_groups.reverse()

        return groups, index_groups

    @staticmethod
    def _init_worker(shared_pos, pos_shape):
        """
        Initialize worker process with shared memory access.

        Sets up global shared memory access for worker processes,
        allowing them to read molecular positions without data copying.

        Args:
            shared_pos: RawArray containing shared position data.
            pos_shape (tuple): Shape tuple for reshaping the shared array.
        """
        global shared_positions
        shared_positions = np.frombuffer(shared_pos, dtype=np.float64).reshape(
            pos_shape
        )

    def _calculate_rmsd(self, idx_pair: Tuple[int, int]) -> float:
        """
        Calculate RMSD efficiently using shared memory.

        Computes RMSD between two molecules by reading their positions
        from shared memory and creating local copies to reduce lock
        contention during computation.

        Args:
            idx_pair (Tuple[int, int]): Pair of molecule indices to compare.

        Returns:
            float: RMSD value between the two molecules, or np.inf if
                   shapes don't match.
        """
        i, j = idx_pair

        # Read from Shared Memory ONCE (No repeated locking)
        pos1 = np.array(shared_positions[i])  # Copying reduces lock contention
        pos2 = np.array(shared_positions[j])

        if pos1.shape != pos2.shape:
            return np.inf

        if self.align_molecules:
            pos1, pos2, _, _, _ = kabsch_align(pos1, pos2)

        return np.sqrt(np.mean(np.sum((pos1 - pos2) ** 2, axis=1)))


class TanimotoSimilarityGrouper(MoleculeGrouper):
    """
    Groups molecules based on fingerprint similarity using Tanimoto coefficient.

    This class supports different fingerprint types and uses connected components
    clustering to group structurally similar molecules.

    Supported fingerprint types:
    - "rdkit": RDKit topological fingerprint (default)
    - "rdk": Legacy RDKit fingerprint
    - "morgan": Morgan (circular) fingerprint (radius=2)
    - "maccs": MACCS keys (166 bits)
    - "atompair": Atom pair fingerprint
    - "torsion": Topological torsion fingerprint
    - "usr": Ultrafast Shape Recognition (3D descriptor)
    - "usrcat": USR with CREDO Atom Types (3D descriptor)
    """

    def __init__(
        self,
        molecules: Iterable[Molecule],
        threshold=None,  # Tanimoto similarity threshold
        num_groups=None,  # Number of groups to create (alternative to threshold)
        num_procs: int = 1,
        fingerprint_type: str = "rdkit",
        use_rdkit_fp: bool = None,  # Legacy support
        label: str = None,  # Label for output files
        ignore_hydrogens: bool = False,
        **kwargs,
    ):
        """
        Initialize Tanimoto similarity-based molecular grouper.

        Args:
            molecules (Iterable[Molecule]): Collection of molecules to group.
            threshold (float): Tanimoto similarity threshold. Defaults to 0.9.
                Ignored if num_groups is specified.
            num_groups (int): Number of groups to create. When specified,
                automatically determines threshold to create this many groups.
            num_procs (int): Number of processes for parallel computation.
            fingerprint_type (str): Type of fingerprint to use.
                Options: "rdkit", "rdk", "morgan", "maccs", "atompair",
                "torsion", "usr", "usrcat". Defaults to "rdkit".
            use_rdkit_fp (bool): Legacy parameter. If True, sets fingerprint_type="rdkit".
                If False, sets fingerprint_type="rdk".
            label (str): Label/name for output files. Defaults to None.
            ignore_hydrogens (bool): Not supported for this grouper. Must be False.
        """
        super().__init__(molecules, num_procs, label=label)

        # Check for unsupported ignore_hydrogens
        if ignore_hydrogens:
            raise ValueError(
                "TanimotoSimilarityGrouper does not support ignore_hydrogens=True. "
                "Molecular fingerprints require complete molecular structure information."
            )

        # Validate that threshold and num_groups are mutually exclusive
        if threshold is not None and num_groups is not None:
            raise ValueError(
                "Cannot specify both threshold (-t) and num_groups (-N). Please use only one."
            )

        if threshold is None and num_groups is None:
            threshold = 0.9
        self.threshold = threshold
        self.num_groups = num_groups

        if use_rdkit_fp is not None:
            self.fingerprint_type = "rdkit" if use_rdkit_fp else "rdk"
        else:
            self.fingerprint_type = fingerprint_type.lower()

        # Convert valid molecules to RDKit format
        self.rdkit_molecules = [
            mol.to_rdkit() for mol in molecules if mol.to_rdkit()
        ]
        self.valid_molecules = [mol for mol in molecules if mol.to_rdkit()]

    def _get_fingerprint(self, rdkit_mol: Chem.Mol) -> Optional[object]:
        """
        Generate a fingerprint for a molecule.

        Args:
            rdkit_mol (Chem.Mol): RDKit molecule object.

        Returns:
            Optional[object]: Molecular fingerprint (BitVect or np.ndarray)
                or None if generation fails.
        """
        try:
            if self.fingerprint_type == "rdkit":
                return GetRDKitFPGenerator().GetFingerprint(rdkit_mol)
            elif self.fingerprint_type == "rdk":
                return Chem.RDKFingerprint(rdkit_mol)
            elif self.fingerprint_type == "morgan":
                return rdMolDescriptors.GetMorganFingerprintAsBitVect(
                    rdkit_mol, 2
                )
            elif self.fingerprint_type == "maccs":
                return rdMolDescriptors.GetMACCSKeysFingerprint(rdkit_mol)
            elif self.fingerprint_type == "atompair":
                return rdMolDescriptors.GetHashedAtomPairFingerprintAsBitVect(
                    rdkit_mol
                )
            elif self.fingerprint_type == "torsion":
                return rdMolDescriptors.GetHashedTopologicalTorsionFingerprintAsBitVect(
                    rdkit_mol
                )
            elif self.fingerprint_type == "usr":
                conf = rdkit_mol.GetConformer()
                return np.array(
                    rdMolDescriptors.GetUSR(rdkit_mol, confId=conf.GetId())
                )

            elif self.fingerprint_type == "usrcat":
                conf = rdkit_mol.GetConformer()
                return np.array(
                    rdMolDescriptors.GetUSRCAT(rdkit_mol, confId=conf.GetId())
                )
            else:
                logger.warning(
                    f"Unknown fingerprint type: {self.fingerprint_type}, using RDKit default."
                )
                return GetRDKitFPGenerator().GetFingerprint(rdkit_mol)
        except Exception as e:
            logger.warning(f"Fingerprint generation failed: {str(e)}")
            return None

    def group(self) -> Tuple[List[List[Molecule]], List[List[int]]]:
        """
        Groups molecules based on Tanimoto similarity clustering.

        Computes fingerprints for all molecules, calculates pairwise
        Tanimoto similarities, and groups molecules using connected
        components clustering.

        Returns:
            Tuple[List[List[Molecule]], List[List[int]]]: Tuple containing:
                - List of molecule groups (each group is a list of molecules)
                - List of index groups (corresponding indices for each group)
        """
        import time

        # Record start time for grouping process
        grouping_start_time = time.time()

        n = len(self.molecules)
        print(
            f"[{self.__class__.__name__}] Starting fingerprint calculation for {n} molecules using {self.fingerprint_type} fingerprints"
        )

        # Compute fingerprints in parallel
        with ThreadPool(self.num_procs) as pool:
            fingerprints = pool.map(
                self._get_fingerprint, self.rdkit_molecules
            )

        # Filter valid fingerprints
        valid_indices = [
            i for i, fp in enumerate(fingerprints) if fp is not None
        ]
        valid_fps = [fingerprints[i] for i in valid_indices]
        num_valid = len(valid_indices)

        if num_valid == 0:
            return [], []  # No valid molecules

        print(
            f"[{self.__class__.__name__}] Computing Tanimoto similarities for {num_valid} valid molecules"
        )

        # Compute similarity matrix
        similarity_matrix = np.zeros((num_valid, num_valid), dtype=np.float32)

        # Check if we are using numpy arrays (USR/USRCAT)
        if valid_fps and isinstance(valid_fps[0], np.ndarray):
            # Calculate Tanimoto for continuous variables (vectors)
            # T(A, B) = (A . B) / (|A|^2 + |B|^2 - A . B)
            fps_array = np.array(valid_fps)

            # Compute pairwise dot products
            dot_products = np.dot(fps_array, fps_array.T)

            # Compute squared norms
            norms_sq = np.diag(dot_products)

            # Compute Tanimoto matrix
            # Denominator: |A|^2 + |B|^2 - A.B
            # Broadcasting: norms_sq[:, None] + norms_sq[None, :] - dot_products
            denominator = norms_sq[:, None] + norms_sq[None, :] - dot_products

            # Avoid division by zero
            denominator[denominator == 0] = 1e-9

            similarity_matrix = dot_products / denominator

        else:
            # Use RDKit DataStructs for BitVects
            pairs = [
                (i, j)
                for i in range(num_valid)
                for j in range(i + 1, num_valid)
            ]

            with ThreadPool(self.num_procs) as pool:
                similarities = pool.starmap(
                    DataStructs.FingerprintSimilarity,
                    [(valid_fps[i], valid_fps[j]) for i, j in pairs],
                )

            # Fill similarity matrix
            for (i, j), sim in zip(pairs, similarities):
                similarity_matrix[i, j] = similarity_matrix[j, i] = sim

        # Choose grouping strategy based on parameters
        if self.num_groups is not None:
            groups, index_groups = self._group_by_num_groups(
                similarity_matrix, valid_indices
            )
        else:
            groups, index_groups = self._group_by_threshold(
                similarity_matrix, valid_indices
            )

        # Calculate total grouping time
        grouping_end_time = time.time()
        grouping_time = grouping_end_time - grouping_start_time

        # Save Tanimoto matrix to group_result folder
        import os

        # Use label for folder name if provided
        if self.label:
            output_dir = f"{self.label}_group_result"
        else:
            output_dir = "group_result"
        os.makedirs(output_dir, exist_ok=True)

        # Create filename based on label, grouper type and threshold/num_groups
        # Format: {label}_{GrouperClass}_T{threshold}.xlsx or {label}_{GrouperClass}_N{num_groups}.xlsx
        label_prefix = f"{self.label}_" if self.label else ""
        if self.num_groups is not None:
            matrix_filename = os.path.join(
                output_dir,
                f"{label_prefix}{self.__class__.__name__}_N{self.num_groups}.xlsx",
            )
        else:
            matrix_filename = os.path.join(
                output_dir,
                f"{label_prefix}{self.__class__.__name__}_T{self.threshold}.xlsx",
            )

        # Save full matrix
        self._save_tanimoto_matrix(
            similarity_matrix, matrix_filename, valid_indices, grouping_time
        )

        return groups, index_groups

    def _group_by_threshold(self, similarity_matrix, valid_indices):
        """Threshold-based grouping for Tanimoto similarity using complete linkage."""
        # Apply threshold and create adjacency matrix
        adj_matrix = similarity_matrix >= self.threshold
        n = len(valid_indices)

        # Complete linkage grouping
        mol_groups, idx_groups = self._complete_linkage_grouping(
            adj_matrix, n, valid_indices
        )

        return mol_groups, idx_groups

    def _complete_linkage_grouping(self, adj_matrix, n, valid_indices):
        """
        Perform complete linkage grouping for Tanimoto similarity (from last to first).

        A structure joins a group only if its similarity to ALL existing members
        of that group is above the threshold.
        """
        assigned = [False] * n
        mol_groups = []
        idx_groups = []

        for i in range(n - 1, -1, -1):
            if assigned[i]:
                continue
            current_group = [i]
            assigned[i] = True

            for j in range(i - 1, -1, -1):
                if assigned[j]:
                    continue
                can_join = all(
                    adj_matrix[j, member] for member in current_group
                )
                if can_join:
                    current_group.append(j)
                    assigned[j] = True

            current_group.sort()
            mol_groups.append(
                [
                    self.valid_molecules[valid_indices[idx]]
                    for idx in current_group
                ]
            )
            idx_groups.append([valid_indices[idx] for idx in current_group])

        mol_groups.reverse()
        idx_groups.reverse()
        return mol_groups, idx_groups

    def _group_by_num_groups(self, similarity_matrix, valid_indices):
        """Automatic grouping to create specified number of groups for Tanimoto similarity."""
        n = len(valid_indices)

        if self.num_groups >= n:
            # If requesting more groups than molecules, each molecule is its own group
            print(
                f"[{self.__class__.__name__}] Requested {self.num_groups} groups but only {n} molecules. Creating {n} groups."
            )
            groups = [[self.valid_molecules[i]] for i in valid_indices]
            index_groups = [[idx] for idx in valid_indices]
            return groups, index_groups

        # Extract similarity values for threshold finding
        similarity_values = []
        for i in range(n):
            for j in range(i + 1, n):
                similarity_values.append(similarity_matrix[i, j])

        # Find appropriate threshold to create desired number of groups
        threshold = self._find_optimal_similarity_threshold(
            similarity_values, similarity_matrix, n
        )

        # Store the auto-determined threshold for summary reporting
        self._auto_threshold = threshold

        print(
            f"[{self.__class__.__name__}] Auto-determined threshold: {threshold:.7f} to create {self.num_groups} groups"
        )

        # Build adjacency matrix with the determined threshold
        adj_matrix = similarity_matrix >= threshold

        # Complete linkage grouping
        groups, index_groups = self._complete_linkage_grouping(
            adj_matrix, n, valid_indices
        )
        actual_groups = len(groups)

        print(
            f"[{self.__class__.__name__}] Created {actual_groups} groups (requested: {self.num_groups})"
        )

        if actual_groups > self.num_groups:
            # If we have too many groups, merge the smallest ones
            groups, index_groups = self._merge_groups_to_target_tanimoto(
                groups, index_groups, adj_matrix
            )

        return groups, index_groups

    def _merge_groups_to_target_tanimoto(
        self, groups, index_groups, adj_matrix
    ):
        """Merge groups to reach target number for Tanimoto grouping."""
        # Convert index_groups to local indices for adj_matrix lookup
        while len(groups) > self.num_groups:
            min_size = float("inf")
            min_idx = 0
            for i, g in enumerate(groups):
                if len(g) < min_size:
                    min_size = len(g)
                    min_idx = i

            # Merge into largest group
            largest_idx = max(
                range(len(groups)),
                key=lambda i: len(groups[i]) if i != min_idx else -1,
            )
            groups[largest_idx].extend(groups[min_idx])
            index_groups[largest_idx].extend(index_groups[min_idx])

            groups.pop(min_idx)
            index_groups.pop(min_idx)

        return groups, index_groups

    def _find_optimal_similarity_threshold(
        self, similarity_values, similarity_matrix, n
    ):
        """Find similarity threshold that creates approximately the desired number of groups using binary search."""
        # Sort similarity values (higher similarity = more similar for Tanimoto)
        sorted_similarities = sorted(
            [sim for sim in similarity_values if not np.isnan(sim)],
            reverse=True,
        )

        if not sorted_similarities:
            return 1.0

        # Binary search for optimal threshold
        low, high = 0, len(sorted_similarities) - 1
        best_threshold = sorted_similarities[0]

        while low <= high:
            mid = (low + high) // 2
            threshold = sorted_similarities[mid]

            # Build adjacency matrix with this threshold
            adj_matrix = similarity_matrix >= threshold

            # Count groups using complete linkage
            num_groups_found = self._count_complete_linkage_groups(
                adj_matrix, n
            )

            if num_groups_found == self.num_groups:
                # Found exact match
                return threshold
            elif num_groups_found > self.num_groups:
                # Too many groups, need lower threshold (less restrictive)
                high = mid - 1
            else:
                # Too few groups, need higher threshold (more restrictive)
                best_threshold = threshold
                low = mid + 1

        return best_threshold

    def _count_complete_linkage_groups(self, adj_matrix, n):
        """Count number of groups using complete linkage algorithm."""
        assigned = [False] * n
        num_groups = 0

        for i in range(n):
            if assigned[i]:
                continue
            current_group = [i]
            assigned[i] = True

            for j in range(i + 1, n):
                if assigned[j]:
                    continue
                can_join = all(
                    adj_matrix[j, member] for member in current_group
                )
                if can_join:
                    current_group.append(j)
                    assigned[j] = True

            num_groups += 1

        return num_groups

    def _merge_smallest_tanimoto_groups(
        self, labels, actual_groups, valid_indices
    ):
        """Merge smallest groups to reach target number of groups."""
        unique_labels = np.unique(labels)
        group_sizes = [
            (label, len(np.where(labels == label)[0]))
            for label in unique_labels
        ]
        group_sizes.sort(key=lambda x: x[1])  # Sort by size

        # Keep largest groups, merge smallest ones into the largest
        groups_to_keep = group_sizes[-self.num_groups :]
        groups_to_merge = group_sizes[: -self.num_groups]

        # Create final groups
        groups = []
        index_groups = []

        # Add the groups we're keeping
        for label, _ in groups_to_keep:
            indices = list(np.where(labels == label)[0])
            groups.append(
                [self.valid_molecules[valid_indices[i]] for i in indices]
            )
            index_groups.append([valid_indices[i] for i in indices])

        # Merge remaining groups into the largest group
        if groups_to_merge:
            merge_indices = []
            for label, _ in groups_to_merge:
                merge_indices.extend(list(np.where(labels == label)[0]))

            # Add to the largest group
            if groups:
                groups[0].extend([self.molecules[i] for i in merge_indices])
                index_groups[0].extend(merge_indices)

        return groups, index_groups

    def _save_tanimoto_matrix(
        self,
        tanimoto_matrix: np.ndarray,
        filename: str,
        valid_indices: List[int],
        grouping_time: float = None,
    ):
        """Save Tanimoto similarity matrix to Excel file with 7 decimal precision."""
        n = len(self.molecules)

        # Change extension to .xlsx if it's .txt
        if filename.endswith(".txt"):
            filename = filename[:-4] + ".xlsx"
        elif not filename.endswith(".xlsx"):
            filename = filename + ".xlsx"

        # Create full matrix with invalid molecules marked as NaN
        full_matrix = np.full((n, n), np.nan)

        # Fill in valid similarities
        for i, idx_i in enumerate(valid_indices):
            for j, idx_j in enumerate(valid_indices):
                full_matrix[idx_i, idx_j] = tanimoto_matrix[i, j]

        # Create DataFrame with simple numeric indices as row and column labels
        row_labels = [str(i + 1) for i in range(n)]
        col_labels = [str(j + 1) for j in range(n)]
        df = pd.DataFrame(full_matrix, index=row_labels, columns=col_labels)

        # Create Excel writer
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            # Write matrix to sheet starting from row 8 to leave room for header info
            df.to_excel(
                writer,
                sheet_name="Tanimoto_Matrix",
                startrow=8,
                float_format="%.7f",
            )

            # Get the worksheet to add header information
            worksheet = writer.sheets["Tanimoto_Matrix"]

            # Add header information
            row = 1
            worksheet[f"A{row}"] = (
                f"Full Tanimoto Similarity Matrix ({n}x{n}) - {self.__class__.__name__}"
            )
            row += 1
            worksheet[f"A{row}"] = f"Fingerprint Type: {self.fingerprint_type}"
            row += 1

            if hasattr(self, "num_groups") and self.num_groups is not None:
                worksheet[f"A{row}"] = (
                    f"Requested Groups (-N): {self.num_groups}"
                )
                row += 1
                if (
                    hasattr(self, "_auto_threshold")
                    and self._auto_threshold is not None
                ):
                    worksheet[f"A{row}"] = (
                        f"Auto-determined Threshold: {self._auto_threshold:.7f}"
                    )
                    row += 1
            else:
                worksheet[f"A{row}"] = f"Threshold: {self.threshold}"
                row += 1

            worksheet[f"A{row}"] = f"Number of Processors: {self.num_procs}"
            row += 1

            if grouping_time is not None:
                worksheet[f"A{row}"] = (
                    f"Grouping Time: {grouping_time:.2f} seconds"
                )
                row += 1

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = min(max_length + 2, 18)
                worksheet.column_dimensions[column_letter].width = (
                    adjusted_width
                )

        logger.info(f"Tanimoto matrix saved to {filename}")

    def __repr__(self):
        if self.num_groups is not None:
            return (
                f"{self.__class__.__name__}(num_groups={self.num_groups}, "
                f"num_procs={self.num_procs}, fingerprint_type={self.fingerprint_type})"
            )
        else:
            return (
                f"{self.__class__.__name__}(threshold={self.threshold}, "
                f"num_procs={self.num_procs}, fingerprint_type={self.fingerprint_type})"
            )

    def calculate_full_tanimoto_matrix(
        self, output_file: Optional[str] = None
    ) -> np.ndarray:
        """
        Calculate the full Tanimoto similarity matrix for all molecule pairs.

        Args:
            output_file (Optional[str]): Optional file path to save the matrix.

        Returns:
            np.ndarray: Full Tanimoto similarity matrix (n x n).
        """
        # Compute fingerprints in parallel
        with ThreadPool(self.num_procs) as pool:
            fingerprints = pool.map(
                self._get_fingerprint, self.rdkit_molecules
            )

        # Filter valid fingerprints
        valid_indices = [
            i for i, fp in enumerate(fingerprints) if fp is not None
        ]
        valid_fps = [fingerprints[i] for i in valid_indices]
        num_valid = len(valid_indices)

        if num_valid == 0:
            return np.array([])  # No valid molecules

        # Compute similarity matrix
        similarity_matrix = np.zeros((num_valid, num_valid), dtype=np.float32)

        # Check if we are using numpy arrays (USR/USRCAT)
        if valid_fps and isinstance(valid_fps[0], np.ndarray):
            # Calculate Tanimoto for continuous variables (vectors)
            # T(A, B) = (A . B) / (|A|^2 + |B|^2 - A . B)
            fps_array = np.array(valid_fps)

            # Compute pairwise dot products
            dot_products = np.dot(fps_array, fps_array.T)

            # Compute squared norms
            norms_sq = np.diag(dot_products)

            # Compute Tanimoto matrix
            # Denominator: |A|^2 + |B|^2 - A.B
            # Broadcasting: norms_sq[:, None] + norms_sq[None, :] - dot_products
            denominator = norms_sq[:, None] + norms_sq[None, :] - dot_products

            # Avoid division by zero
            denominator[denominator == 0] = 1e-9

            similarity_matrix = dot_products / denominator

        else:
            # Use RDKit DataStructs for BitVects
            pairs = [
                (i, j)
                for i in range(num_valid)
                for j in range(i + 1, num_valid)
            ]

            with ThreadPool(self.num_procs) as pool:
                similarities = pool.starmap(
                    DataStructs.FingerprintSimilarity,
                    [(valid_fps[i], valid_fps[j]) for i, j in pairs],
                )

            # Fill similarity matrix
            for (i, j), sim in zip(pairs, similarities):
                similarity_matrix[i, j] = similarity_matrix[j, i] = sim

        # Create full matrix with invalid molecules marked as NaN
        n = len(self.molecules)
        full_matrix = np.full((n, n), np.nan)

        # Fill in valid similarities
        for i, idx_i in enumerate(valid_indices):
            for j, idx_j in enumerate(valid_indices):
                full_matrix[idx_i, idx_j] = similarity_matrix[i, j]

        # Save to file if requested
        if output_file:
            self._save_tanimoto_matrix(
                similarity_matrix, output_file, valid_indices, None
            )

        return full_matrix


class RDKitIsomorphismGrouper(MoleculeGrouper):
    """
    Group molecules using RDKit hashing with optional isomorphism checks.

    First clusters molecules by RDKit molecular hash (choice depends on
    options), then optionally verifies equivalence using InChIKey equality
    as a lightweight isomorphism proxy. This can be computationally
    expensive for large sets.

    Hashing choices (see `_get_mol_hash`):
    - If `use_tautomers` is True: `rdMolHash.HashFunction.Tautomer`.
    - Else if `use_stereochemistry` is True: `rdMolHash.HashFunction.AnonymousGraph`.
    - Else: `rdMolHash.HashFunction.MolFormula`.

    Attributes:
        molecules (Iterable[Molecule]): Inherited; collection of molecules to
            group.
        num_procs (int): Inherited; number of worker processes.
        use_stereochemistry (bool): Whether to consider stereochemistry.
        use_tautomers (bool): Whether to consider tautomeric forms.
    """

    def __init__(
        self,
        molecules: Iterable[Molecule],
        num_procs: int = 1,
        use_stereochemistry: bool = True,
        use_tautomers: bool = False,
        label: str = None,
        ignore_hydrogens: bool = False,
        **kwargs,
    ):
        """
        Initialize RDKit isomorphism-based molecular grouper.

        Args:
            molecules (Iterable[Molecule]): Collection of molecules to group.
            num_procs (int): Number of processes for parallel computation.
            use_stereochemistry (bool): Whether to consider stereochemical
                differences in grouping. Defaults to True.
            use_tautomers (bool): Whether to consider tautomeric forms
                as equivalent. Defaults to False.
            label (str): Label/name for output files. Defaults to None.
            ignore_hydrogens (bool): Not supported for this grouper. Must be False.
        """
        super().__init__(molecules, num_procs, label=label)

        # Check for unsupported ignore_hydrogens
        if ignore_hydrogens:
            raise ValueError(
                "RDKitIsomorphismGrouper does not support ignore_hydrogens=True. "
                "RDKit molecular hashing requires complete molecular structure information."
            )

        self.use_stereochemistry = use_stereochemistry
        self.use_tautomers = use_tautomers

    def _get_mol_hash(self, mol: Molecule) -> Optional[str]:
        """
        Generate canonical hash for molecular structure identification.

        Creates a canonical hash string for the molecule using RDKit's
        hashing functions. Hash type depends on stereochemistry and
        tautomer configuration settings.

        Args:
            mol (Molecule): Molecule to generate hash for.

        Returns:
            Optional[str]: Canonical hash string, or None if generation fails.
        """
        try:
            rdkit_mol = mol.to_rdkit()
            if not rdkit_mol:
                return None
            # Choose hashing function based on requirements
            hash_func = (
                rdMolHash.HashFunction.Tautomer
                if self.use_tautomers
                else (
                    rdMolHash.HashFunction.AnonymousGraph
                    if self.use_stereochemistry
                    else rdMolHash.HashFunction.MolFormula
                )
            )
            return rdMolHash.MolHash(rdkit_mol, hash_func)
        except Exception as e:
            logger.warning(f"Hash generation failed: {str(e)}")
            return None

    def group(self) -> Tuple[List[List[Molecule]], List[List[int]]]:
        """
        Group molecules using structural isomorphism detection.

        Uses RDKit molecular hashing for initial grouping, followed by
        detailed isomorphism checks when stereochemistry or tautomer
        considerations are enabled.

        Returns:
            Tuple[List[List[Molecule]], List[List[int]]]: Tuple containing:
                - List of molecule groups (each group is a list of molecules)
                - List of index groups (corresponding indices for each group)
        """
        import time

        grouping_start_time = time.time()

        with multiprocessing.Pool(self.num_procs) as pool:
            hashes = pool.map(self._get_mol_hash, self.molecules)

        n = len(self.molecules)

        # Build adjacency matrix (1 = isomorphic, 0 = not)
        adj_matrix = np.zeros((n, n), dtype=bool)
        for i in range(n):
            for j in range(i, n):
                same_hash = hashes[i] == hashes[j]
                if same_hash:
                    # Verify isomorphism if needed
                    if self.use_stereochemistry or self.use_tautomers:
                        is_iso = self._check_isomorphism(
                            self.molecules[i], self.molecules[j]
                        )
                    else:
                        is_iso = True
                    if is_iso:
                        adj_matrix[i, j] = adj_matrix[j, i] = True

        groups: List[List[Molecule]] = []
        indices = list(range(n))
        index_groups: List[List[int]] = []

        while indices:
            pivot_idx = indices[0]
            # Find all molecules isomorphic to pivot
            matches = [i for i in indices if adj_matrix[pivot_idx, i]]

            groups.append([self.molecules[i] for i in matches])
            index_groups.append(matches)
            indices = [i for i in indices if i not in matches]

        grouping_time = time.time() - grouping_start_time

        # Save grouping results to Excel with matrix
        self._save_groups_to_excel(adj_matrix, index_groups, grouping_time)

        # Cache the results to avoid recomputation
        self._cached_groups = groups
        self._cached_group_indices = index_groups

        return groups, index_groups

    def _save_groups_to_excel(
        self, adj_matrix, idx_groups, grouping_time=None
    ):
        """Save isomorphism grouping results to Excel file with 2 sheets."""
        if self.label:
            output_dir = f"{self.label}_group_result"
        else:
            output_dir = "group_result"
        os.makedirs(output_dir, exist_ok=True)

        # Filename format: {label}_{GrouperClass}.xlsx
        label_prefix = f"{self.label}_" if self.label else ""
        filename = os.path.join(
            output_dir, f"{label_prefix}{self.__class__.__name__}.xlsx"
        )

        n = adj_matrix.shape[0]

        # Create DataFrame with simple numeric indices
        row_labels = [str(i + 1) for i in range(n)]
        col_labels = [str(j + 1) for j in range(n)]

        # Convert boolean matrix to int (1/0) for display
        matrix_display = adj_matrix.astype(int)
        df = pd.DataFrame(matrix_display, index=row_labels, columns=col_labels)

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            # Sheet 1: Isomorphism_Matrix
            df.to_excel(
                writer,
                sheet_name="Isomorphism_Matrix",
                startrow=10,
            )

            worksheet = writer.sheets["Isomorphism_Matrix"]

            # Add header information
            row = 1
            worksheet[f"A{row}"] = (
                f"Isomorphism Grouping Matrix ({n}x{n}) - {self.__class__.__name__}"
            )
            row += 1
            worksheet[f"A{row}"] = f"Label: {self.label or 'N/A'}"
            row += 1
            worksheet[f"A{row}"] = f"Total Molecules: {n}"
            row += 1
            worksheet[f"A{row}"] = f"Groups Found: {len(idx_groups)}"
            row += 1
            worksheet[f"A{row}"] = (
                f"Use Stereochemistry: {self.use_stereochemistry}"
            )
            row += 1
            worksheet[f"A{row}"] = f"Use Tautomers: {self.use_tautomers}"
            row += 1

            if grouping_time is not None:
                worksheet[f"A{row}"] = (
                    f"Grouping Time: {grouping_time:.2f} seconds"
                )
                row += 1

            worksheet[f"A{row}"] = (
                "1 = Isomorphic (same structure), 0 = Not isomorphic"
            )
            row += 1

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = min(max_length + 2, 18)
                worksheet.column_dimensions[column_letter].width = (
                    adjusted_width
                )

            # Sheet 2: Groups
            ws_groups = writer.book.create_sheet("Groups")

            # Write header
            ws_groups["A1"] = "Group"
            ws_groups["B1"] = "Members"
            ws_groups["C1"] = "Indices"

            # Write data
            for i, indices in enumerate(idx_groups):
                ws_groups[f"A{i+2}"] = i + 1
                ws_groups[f"B{i+2}"] = len(indices)
                ws_groups[f"C{i+2}"] = str([idx + 1 for idx in indices])

            # Auto-adjust column widths
            ws_groups.column_dimensions["A"].width = 8
            ws_groups.column_dimensions["B"].width = 10
            ws_groups.column_dimensions["C"].width = 50

        logger.info(f"Isomorphism grouping results saved to {filename}")

    def _check_isomorphism(self, mol1: Molecule, mol2: Molecule) -> bool:
        """
        Check equivalence via InChIKey (isomorphism proxy).

        Compares the RDKit InChIKey strings of both molecules. This is a
        lightweight proxy for isomorphism when stereochemistry or tautomer
        checks are enabled; it does not perform an explicit graph
        isomorphism test.

        Args:
            mol1 (Molecule): First molecule to compare.
            mol2 (Molecule): Second molecule to compare.

        Returns:
            bool: True if InChIKeys match; False otherwise (including on failure).
        """
        try:
            return Chem.MolToInchiKey(mol1.to_rdkit()) == Chem.MolToInchiKey(
                mol2.to_rdkit()
            )
        except Exception as e:
            logger.warning(f"Isomorphism check failed: {str(e)}")
            return False


class FormulaGrouper(MoleculeGrouper):
    """
    Group molecules by chemical formula.

    Groups molecules based solely on their chemical formula composition,
    making it suitable when elemental composition is the primary concern.
    Ideal for initial filtering and broad chemical classification.
    """

    def __init__(
        self,
        molecules: Iterable[Molecule],
        num_procs: int = 1,
        label: str = None,
        ignore_hydrogens: bool = False,
        **kwargs,
    ):
        """
        Initialize formula-based molecular grouper.

        Args:
            molecules (Iterable[Molecule]): Collection of molecules to group.
            num_procs (int): Number of processes for parallel computation.
            label (str): Label/name for output files. Defaults to None.
            ignore_hydrogens (bool): Not supported for this grouper. Must be False.
        """
        super().__init__(molecules, num_procs, label=label)

        # Check for unsupported ignore_hydrogens
        if ignore_hydrogens:
            raise ValueError(
                "FormulaGrouper does not support ignore_hydrogens=True. "
                "Chemical formula grouping requires complete molecular composition."
            )

    def group(self):
        """
        Group molecules by chemical formula composition.

        Creates groups based on identical chemical formulas, regardless
        of structural or stereochemical differences. Each group contains
        molecules with the same elemental composition.

        Returns:
            Tuple[List[List[Molecule]], List[List[int]]]: Tuple containing:
                - List of molecule groups (each group is a list of molecules)
                - List of index groups (corresponding indices for each group)
        """
        import time

        grouping_start_time = time.time()

        # Get formula for each molecule
        mol_formulas = [mol.get_chemical_formula() for mol in self.molecules]
        n = len(self.molecules)

        # Build adjacency matrix (1 = same formula, 0 = different)
        adj_matrix = np.zeros((n, n), dtype=bool)
        for i in range(n):
            for j in range(i, n):
                if mol_formulas[i] == mol_formulas[j]:
                    adj_matrix[i, j] = adj_matrix[j, i] = True

        # Group by formula
        formula_groups = {}
        for idx, formula in enumerate(mol_formulas):
            if formula not in formula_groups:
                formula_groups[formula] = []
            formula_groups[formula].append(idx)

        mol_groups = []
        idx_groups = []
        formulas = []
        for formula, indices in formula_groups.items():
            mol_groups.append([self.molecules[i] for i in indices])
            idx_groups.append(indices)
            formulas.append(formula)

        grouping_time = time.time() - grouping_start_time

        # Save formula grouping results to Excel with matrix
        self._save_formula_to_excel(
            adj_matrix, formulas, idx_groups, grouping_time
        )

        return mol_groups, idx_groups

    def _save_formula_to_excel(
        self, adj_matrix, formulas, idx_groups, grouping_time=None
    ):
        """Save formula grouping results to Excel file with 2 sheets."""
        if self.label:
            output_dir = f"{self.label}_group_result"
        else:
            output_dir = "group_result"
        os.makedirs(output_dir, exist_ok=True)

        # Filename format: {label}_{GrouperClass}.xlsx
        label_prefix = f"{self.label}_" if self.label else ""
        filename = os.path.join(
            output_dir, f"{label_prefix}{self.__class__.__name__}.xlsx"
        )

        n = adj_matrix.shape[0]

        # Create DataFrame with simple numeric indices
        row_labels = [str(i + 1) for i in range(n)]
        col_labels = [str(j + 1) for j in range(n)]

        # Convert boolean matrix to int (1/0) for display
        matrix_display = adj_matrix.astype(int)
        df = pd.DataFrame(matrix_display, index=row_labels, columns=col_labels)

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            # Sheet 1: Formula_Matrix
            df.to_excel(
                writer,
                sheet_name="Formula_Matrix",
                startrow=10 + len(formulas),
            )

            worksheet = writer.sheets["Formula_Matrix"]

            # Add header information
            row = 1
            worksheet[f"A{row}"] = (
                f"Formula Grouping Matrix ({n}x{n}) - {self.__class__.__name__}"
            )
            row += 1
            worksheet[f"A{row}"] = f"Label: {self.label or 'N/A'}"
            row += 1
            worksheet[f"A{row}"] = f"Total Molecules: {n}"
            row += 1
            worksheet[f"A{row}"] = f"Groups Found: {len(formulas)}"
            row += 1

            if grouping_time is not None:
                worksheet[f"A{row}"] = (
                    f"Grouping Time: {grouping_time:.2f} seconds"
                )
                row += 1

            worksheet[f"A{row}"] = "1 = Same formula, 0 = Different formula"
            row += 2

            # List all found formulas
            worksheet[f"A{row}"] = "Found Formulas:"
            row += 1
            for i, formula in enumerate(formulas):
                worksheet[f"A{row}"] = f"Formula {i + 1}: {formula}"
                row += 1

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = min(max_length + 2, 18)
                worksheet.column_dimensions[column_letter].width = (
                    adjusted_width
                )

            # Sheet 2: Groups (with 4 columns: Group, Formula, Members, Indices)
            ws_groups = writer.book.create_sheet("Groups")

            # Write header
            ws_groups["A1"] = "Group"
            ws_groups["B1"] = "Formula"
            ws_groups["C1"] = "Members"
            ws_groups["D1"] = "Indices"

            # Write data
            for i, (formula, indices) in enumerate(zip(formulas, idx_groups)):
                ws_groups[f"A{i+2}"] = i + 1
                ws_groups[f"B{i+2}"] = formula
                ws_groups[f"C{i+2}"] = len(indices)
                ws_groups[f"D{i+2}"] = str([idx + 1 for idx in indices])

            # Auto-adjust column widths
            ws_groups.column_dimensions["A"].width = 8
            ws_groups.column_dimensions["B"].width = 20
            ws_groups.column_dimensions["C"].width = 10
            ws_groups.column_dimensions["D"].width = 50

        logger.info(f"Formula grouping results saved to {filename}")


class ConnectivityGrouper(MoleculeGrouper):
    """
    Group molecules based on molecular connectivity (graph isomorphism).

    Groups molecules by analyzing their bond connectivity patterns using
    graph isomorphism. Efficient for recognizing similar bond arrangements
    in large datasets regardless of 3D spatial configuration.

    Attributes:
        molecules (Iterable[Molecule]): Inherited; collection of molecules to
            group.
        num_procs (int): Inherited; number of worker processes.
        threshold (float): Buffer for bond cutoff distance.
        adjust_H (bool): Whether to adjust hydrogen bond detection.
        ignore_hydrogens (bool): Whether to exclude hydrogen atoms from comparison.
    """

    def __init__(
        self,
        molecules: Iterable[Molecule],
        num_procs: int = 1,
        threshold=None,  # Buffer for bond cutoff
        adjust_H: bool = True,
        ignore_hydrogens: bool = False,
        label: str = None,
        **kwargs,
    ):
        """
        Initialize connectivity-based molecular grouper.

        Args:
            molecules (Iterable[Molecule]): Collection of molecules to group.
            num_procs (int): Number of processes for parallel computation.
            threshold (float): Buffer for bond cutoff distance. Defaults to 0.0.
            adjust_H (bool): Whether to adjust hydrogen bond detection.
                Defaults to True.
            ignore_hydrogens (bool): Whether to exclude hydrogen atoms from
                graph comparison. Defaults to False.
            label (str): Label/name for output files. Defaults to None.
        """
        super().__init__(molecules, num_procs, label=label)
        if threshold is None:
            threshold = 0.0
        self.threshold = threshold  # Buffer for bond cutoff
        self.adjust_H = adjust_H
        self.ignore_hydrogens = ignore_hydrogens

    def _are_isomorphic(self, g1: nx.Graph, g2: nx.Graph) -> bool:
        """
        Check if two molecular graphs are isomorphic (NetworkX).

        Uses `networkx.is_isomorphic` with attribute-aware matching:
        - Nodes must have equal `element` values.
        - Edges must have equal `bond_order` values.

        Args:
            g1 (nx.Graph): First molecular graph.
            g2 (nx.Graph): Second molecular graph.

        Returns:
            bool: True if graphs are isomorphic, False otherwise.
        """
        return nx.is_isomorphic(
            g1,
            g2,
            node_match=lambda a, b: a["element"] == b["element"],
            edge_match=lambda a, b: a["bond_order"] == b["bond_order"],
        )

    def _check_isomorphism(
        self, idx_pair: Tuple[int, int]
    ) -> Tuple[int, int, bool]:
        """
        Check graph isomorphism between two molecules for multiprocessing.

        Multiprocessing-compatible function that checks whether two
        molecular graphs are isomorphic based on their connectivity
        patterns and atomic properties.

        Args:
            idx_pair (Tuple[int, int]): Pair of molecule indices to compare.

        Returns:
            Tuple[int, int, bool]: Original indices and isomorphism result.
        """
        i, j = idx_pair
        return i, j, self._are_isomorphic(self.graphs[i], self.graphs[j])

    def group(self) -> Tuple[List[List[Molecule]], List[List[int]]]:
        """
        Group molecules by connectivity using parallel isomorphism checks.

        Converts molecules to graph representations and performs pairwise
        isomorphism checks in parallel. Uses connected components clustering
        to identify groups of structurally equivalent molecules.

        Returns:
            Tuple[List[List[Molecule]], List[List[int]]]: Tuple containing:
                - List of molecule groups (each group is a list of molecules)
                - List of index groups (corresponding indices for each group)
        """
        import time

        grouping_start_time = time.time()

        n = len(self.molecules)

        # Convert molecules to graphs in parallel
        with multiprocessing.Pool(self.num_procs) as pool:
            self.graphs = pool.starmap(
                to_graph_wrapper,
                [
                    (mol, self.threshold, self.adjust_H)
                    for mol in self.molecules
                ],
            )

        # Compute pairwise isomorphism in parallel
        indices = [(i, j) for i in range(n) for j in range(i + 1, n)]
        with multiprocessing.Pool(self.num_procs) as pool:
            isomorphic_pairs = pool.map(self._check_isomorphism, indices)

        # Build adjacency matrix (isomorphism matrix: 1 = isomorphic, 0 = not)
        adj_matrix = np.zeros((n, n), dtype=bool)
        for i, j, is_iso in isomorphic_pairs:
            if is_iso:
                adj_matrix[i, j] = adj_matrix[j, i] = True

        # Complete linkage grouping
        groups, index_groups = self._complete_linkage_grouping(adj_matrix, n)

        # Calculate grouping time
        grouping_time = time.time() - grouping_start_time

        # Save isomorphism matrix and groups to Excel
        self._save_connectivity_matrix(adj_matrix, index_groups, grouping_time)

        # Cache the results
        self._cached_groups = groups
        self._cached_group_indices = index_groups

        return groups, index_groups

    def _save_connectivity_matrix(
        self,
        adj_matrix: np.ndarray,
        index_groups: List[List[int]],
        grouping_time: float = None,
    ):
        """Save connectivity (isomorphism) matrix to Excel file."""
        n = adj_matrix.shape[0]

        # Create output directory
        if self.label:
            output_dir = f"{self.label}_group_result"
        else:
            output_dir = "group_result"
        os.makedirs(output_dir, exist_ok=True)

        # Create filename
        label_prefix = f"{self.label}_" if self.label else ""
        filename = os.path.join(
            output_dir,
            f"{label_prefix}{self.__class__.__name__}.xlsx",
        )

        # Create DataFrame with simple numeric indices
        row_labels = [str(i + 1) for i in range(n)]
        col_labels = [str(j + 1) for j in range(n)]

        # Convert boolean matrix to int (1/0) for display
        matrix_display = adj_matrix.astype(int)
        df = pd.DataFrame(matrix_display, index=row_labels, columns=col_labels)

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            # Sheet 1: Connectivity_Matrix
            df.to_excel(
                writer,
                sheet_name="Connectivity_Matrix",
                startrow=8,
            )

            worksheet = writer.sheets["Connectivity_Matrix"]

            # Add header information
            row = 1
            worksheet[f"A{row}"] = (
                f"Connectivity (Isomorphism) Matrix ({n}x{n}) - {self.__class__.__name__}"
            )
            row += 1
            worksheet[f"A{row}"] = f"Bond Cutoff Buffer: {self.threshold}"
            row += 1
            worksheet[f"A{row}"] = f"Adjust H: {self.adjust_H}"
            row += 1
            worksheet[f"A{row}"] = f"Ignore Hydrogens: {self.ignore_hydrogens}"
            row += 1

            if grouping_time is not None:
                worksheet[f"A{row}"] = (
                    f"Grouping Time: {grouping_time:.2f} seconds"
                )
                row += 1

            worksheet[f"A{row}"] = (
                "1 = Isomorphic (same connectivity), 0 = Not isomorphic"
            )
            row += 1

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = min(max_length + 2, 18)
                worksheet.column_dimensions[column_letter].width = (
                    adjusted_width
                )

            # Sheet 2: Groups
            groups_data = []
            for i, indices in enumerate(index_groups):
                groups_data.append(
                    {
                        "Group": i + 1,
                        "Members": len(indices),
                        "Indices": str([idx + 1 for idx in indices]),
                    }
                )

            groups_df = pd.DataFrame(groups_data)
            groups_df.to_excel(writer, sheet_name="Groups", index=False)

            ws_groups = writer.sheets["Groups"]
            ws_groups.column_dimensions["A"].width = 8
            ws_groups.column_dimensions["B"].width = 10
            ws_groups.column_dimensions["C"].width = 50

        logger.info(f"Connectivity matrix saved to {filename}")

    def _complete_linkage_grouping(self, adj_matrix, n):
        """
        Perform complete linkage grouping for connectivity (from last to first).

        A molecule joins a group only if it is isomorphic to ALL existing members.
        """
        assigned = [False] * n
        groups = []
        index_groups = []

        for i in range(n - 1, -1, -1):
            if assigned[i]:
                continue
            current_group = [i]
            assigned[i] = True

            for j in range(i - 1, -1, -1):
                if assigned[j]:
                    continue
                can_join = all(
                    adj_matrix[j, member] for member in current_group
                )
                if can_join:
                    current_group.append(j)
                    assigned[j] = True

            current_group.sort()
            groups.append([self.molecules[idx] for idx in current_group])
            index_groups.append(current_group)

        groups.reverse()
        index_groups.reverse()
        return groups, index_groups


class ConnectivityGrouperSharedMemory(MoleculeGrouper):
    """
    Group molecules based on molecular connectivity using shared memory.

    Optimized version of ConnectivityGrouper that uses shared memory
    for storing molecular graph data to reduce memory overhead in
    multiprocessing scenarios. Particularly useful for large molecular
    datasets where memory efficiency is critical.

    Attributes:
        molecules (Iterable[Molecule]): Inherited; collection of molecules to
            group.
        num_procs (int): Inherited; number of worker processes.
        threshold (float): Buffer for bond cutoff distance.
        adjust_H (bool): Whether to adjust hydrogen bond detection.
        ignore_hydrogens (bool): Whether to exclude hydrogen atoms from comparison.
    """

    def __init__(
        self,
        molecules: Iterable[Molecule],
        num_procs: int = 1,
        threshold: float = 0.0,  # Buffer for bond cutoff
        adjust_H: bool = True,
        ignore_hydrogens: bool = False,
        label: str = None,
        **kwargs,
    ):
        """
        Initialize connectivity grouper with shared memory optimization.

        Args:
            molecules (Iterable[Molecule]): Collection of molecules to group.
            num_procs (int): Number of processes for parallel computation.
            threshold (float): Buffer for bond cutoff distance. Defaults to 0.0.
            adjust_H (bool): Whether to adjust hydrogen bond detection.
                Defaults to True.
            ignore_hydrogens (bool): Whether to exclude hydrogen atoms from
                graph comparison. Defaults to False.
            label (str): Label/name for output files. Defaults to None.
        """
        super().__init__(molecules, num_procs, label=label)
        self.threshold = threshold
        self.adjust_H = adjust_H
        self.ignore_hydrogens = ignore_hydrogens

    def _are_isomorphic(self, g1: nx.Graph, g2: nx.Graph) -> bool:
        """
        Check if two molecular graphs are isomorphic (NetworkX).

        Uses `networkx.is_isomorphic` with attribute-aware matching:
        - Nodes must have equal `element` values.
        - Edges must have equal `bond_order` values.

        Args:
            g1 (nx.Graph): First molecular graph.
            g2 (nx.Graph): Second molecular graph.

        Returns:
            bool: True if graphs are isomorphic, False otherwise.
        """
        return nx.is_isomorphic(
            g1,
            g2,
            node_match=lambda a, b: a["element"] == b["element"],
            edge_match=lambda a, b: a["bond_order"] == b["bond_order"],
        )

    def _check_isomorphism(
        self, pivot_graph_bytes, mol_graph_bytes, idx: int
    ) -> Tuple[int, bool]:
        """
        Check equivalence for multiprocessing using serialized graphs.

        Deserializes NetworkX graphs from bytes and tests connectivity
        isomorphism against a pivot graph.

        Args:
            pivot_graph_bytes (bytes): Pickled pivot molecular graph.
            mol_graph_bytes (bytes): Pickled molecular graph to compare.
            idx (int): Index of the molecule being compared.

        Returns:
            Tuple[int, bool]: (idx, is_isomorphic) result.
        """
        pivot_graph = pickle.loads(pivot_graph_bytes)
        mol_graph = pickle.loads(mol_graph_bytes)
        return idx, self._are_isomorphic(pivot_graph, mol_graph)

    def _convert_to_graphs(self):
        """
        Convert molecules to graphs and store in shared memory.

        Converts all molecules to NetworkX graph representations in parallel,
        then serializes them using pickle and stores in shared memory for
        efficient access by worker processes.

        Returns:
            Tuple[shared_memory.SharedMemory, tuple, np.dtype]: Shared memory
            handle, array shape, and dtype for reconstructing the object array.
        """
        with multiprocessing.Pool(self.num_procs) as pool:
            graphs = pool.starmap(
                to_graph_wrapper,
                [
                    (mol, self.threshold, self.adjust_H)
                    for mol in self.molecules
                ],
            )

        # Serialize graphs using pickle and store in shared memory
        graph_bytes = [pickle.dumps(graph) for graph in graphs]

        # Store in shared NumPy array
        shared_array = np.array(graph_bytes, dtype=object)
        shm = shared_memory.SharedMemory(create=True, size=shared_array.nbytes)
        np.ndarray(
            shared_array.shape, dtype=shared_array.dtype, buffer=shm.buf
        )[:] = shared_array[:]

        return shm, shared_array.shape, shared_array.dtype

    def group(self) -> Tuple[List[List[Molecule]], List[List[int]]]:
        """
        Group molecules using connectivity with shared memory optimization.

        Groups molecules based on molecular connectivity using multiprocessing
        and shared memory for graph storage. Uses iterative comparison with
        a pivot molecule approach to identify structurally equivalent groups.

        Returns:
            Tuple[List[List[Molecule]], List[List[int]]]: Tuple containing:
                - List of molecule groups (each group is a list of molecules)
                - List of index groups (corresponding indices for each group)
        """
        groups = []
        shm, shape, dtype = self._convert_to_graphs()
        graphs = np.ndarray(shape, dtype=dtype, buffer=shm.buf)

        remaining = list(enumerate(zip(self.molecules, graphs)))

        while remaining:
            pivot_idx, (pivot_mol, pivot_graph_bytes) = remaining.pop(0)

            # Parallel isomorphism check
            results = Parallel(n_jobs=self.num_procs, backend="loky")(
                delayed(self._check_isomorphism)(
                    pivot_graph_bytes, g_bytes, idx
                )
                for idx, (_, g_bytes) in remaining
            )

            # Collect isomorphic molecules
            to_remove = {idx for idx, is_iso in results if is_iso}
            current_group = [pivot_mol] + [
                mol for idx, (mol, _) in remaining if idx in to_remove
            ]
            current_indices = [pivot_idx] + [
                idx for idx, _ in remaining if idx in to_remove
            ]

            remaining = [
                (idx, (mol, g_bytes))
                for idx, (mol, g_bytes) in remaining
                if idx not in to_remove
            ]
            groups.append((current_group, current_indices))

        shm.close()
        shm.unlink()  # Free shared memory

        mol_groups = [g[0] for g in groups]
        idx_groups = [g[1] for g in groups]
        return mol_groups, idx_groups


class TorsionFingerprintGrouper(MoleculeGrouper):
    """
    Groups molecule conformers based on RDKit Torsion Fingerprint Deviation (TFD).

    Implementation based on Schulz-Gasch et al., JCIM, 52, 1499-1512 (2012).
    Uses RDKit's TorsionFingerprints.GetTFDBetweenConformers() to analyze different
    conformations of the same molecule and groups similar conformers.

    This grouper follows the same workflow as RMSDGrouper but uses torsion angle
    patterns instead of atomic positions for similarity assessment.

    Attributes:
        molecules (Iterable[Molecule]): Collection of molecules to group.
        threshold (float): TFD threshold for grouping (lower values = more similar).
        num_groups (int): Number of groups to create (alternative to threshold).
        num_procs (int): Number of processes for parallel computation.
        use_weights (bool): Whether to use torsion weights in TFD calculation.
        max_dev (str): Normalization method ('equal' or 'spec').
        symm_radius (int): Radius for calculating atom invariants.
        ignore_colinear_bonds (bool): Whether to ignore single bonds adjacent to triple bonds.
    """

    def __init__(
        self,
        molecules: Iterable[Molecule],
        threshold: float = None,
        num_groups: int = None,
        num_procs: int = 1,
        use_weights: bool = True,
        max_dev: str = "equal",
        symm_radius: int = 2,
        ignore_colinear_bonds: bool = True,
        ignore_hydrogens: bool = False,
        label: str = None,  # Label for output files
        **kwargs,
    ):
        """
        Initialize TFD-based conformer grouper.

        Args:
            molecules (Iterable[Molecule]): Collection of molecule conformers to group.
            threshold (float): TFD threshold for grouping. Lower values indicate more
                similar torsion patterns. Defaults to 0.1. Ignored if num_groups is specified.
            num_groups (int): Number of groups to create. When specified,
                automatically determines threshold to create this many groups.
            num_procs (int): Number of processes for parallel computation.
            use_weights (bool): Whether to use torsion weights in TFD calculation. Defaults to True.
            max_dev (str): Normalization method:
                - 'equal': all torsions normalized using 180.0 (default)
                - 'spec': each torsion normalized using specific maximal deviation
            symm_radius (int): Radius for calculating atom invariants. Defaults to 2.
            ignore_colinear_bonds (bool): If True, single bonds adjacent to triple bonds
                are ignored. Defaults to True.
            ignore_hydrogens (bool): Whether to exclude hydrogen atoms from TFD calculation.
                Defaults to False.
            label (str): Label/name for output files. Defaults to None.
        """
        super().__init__(molecules, num_procs, label=label)

        # Validate that threshold and num_groups are mutually exclusive
        if threshold is not None and num_groups is not None:
            raise ValueError(
                "Cannot specify both threshold (-t) and num_groups (-N). Please use only one."
            )

        if threshold is None and num_groups is None:
            threshold = 0.1  # TFD threshold (lower = more similar)
        self.threshold = threshold
        self.num_groups = num_groups
        self.use_weights = use_weights
        self.max_dev = max_dev
        self.symm_radius = symm_radius
        self.ignore_colinear_bonds = ignore_colinear_bonds
        self.ignore_hydrogens = ignore_hydrogens

        # Prepare molecules for TFD analysis
        self.molecules = list(molecules)  # Convert to list for indexing
        self._prepare_conformer_molecule()

    def _prepare_conformer_molecule(self):
        """
        Prepare a single RDKit molecule with multiple conformers from input molecules.

        This assumes all input molecules are conformers of the same chemical structure.
        Creates one RDKit molecule object with multiple conformer IDs.
        If ignore_hydrogens is True, removes hydrogen atoms from the molecule.
        """
        if not self.molecules:
            self.rdkit_mol = None
            self.valid_conformer_ids = []
            return

        # Use the first molecule as the base structure
        base_mol = self.molecules[0]
        self.rdkit_mol = base_mol.to_rdkit()

        if self.rdkit_mol is None:
            self.valid_conformer_ids = []
            return

        # Remove hydrogens if ignore_hydrogens is True
        if self.ignore_hydrogens:
            self.rdkit_mol = Chem.RemoveHs(self.rdkit_mol)
            logger.info("Removed hydrogen atoms for TFD calculation")

        # Clear existing conformers and add all input molecules as conformers
        self.rdkit_mol.RemoveAllConformers()
        self.valid_conformer_ids = []

        # Get heavy atom indices for filtering positions if ignore_hydrogens
        if self.ignore_hydrogens:
            heavy_atom_indices = [
                i
                for i, sym in enumerate(base_mol.chemical_symbols)
                if sym != "H"
            ]

        for i, mol in enumerate(self.molecules):
            try:
                # Get positions, filtering out hydrogens if needed
                if self.ignore_hydrogens:
                    positions = mol.positions[heavy_atom_indices]
                    num_atoms = len(heavy_atom_indices)
                else:
                    positions = mol.positions
                    num_atoms = mol.num_atoms

                # Convert molecule positions to RDKit conformer
                conf = Chem.Conformer(num_atoms)
                for atom_idx, pos in enumerate(positions):
                    conf.SetAtomPosition(atom_idx, pos)

                # Add conformer to the molecule
                conf_id = self.rdkit_mol.AddConformer(conf, assignId=True)
                self.valid_conformer_ids.append(conf_id)

            except Exception as e:
                logger.warning(f"Failed to add conformer {i}: {str(e)}")
                continue

        logger.info(
            f"Prepared molecule with {len(self.valid_conformer_ids)} valid conformers"
            + (" (hydrogens ignored)" if self.ignore_hydrogens else "")
        )

    def _calculate_tfd(self, conf_pair: Tuple[int, int]) -> float:
        """
        Calculate TFD between two conformers.

        Args:
            conf_pair (Tuple[int, int]): Indices of conformers to compare.

        Returns:
            float: TFD value between the conformers.
        """
        if self.rdkit_mol is None or len(self.valid_conformer_ids) == 0:
            return float("inf")

        i, j = conf_pair

        try:
            conf_id1 = self.valid_conformer_ids[i]
            conf_id2 = self.valid_conformer_ids[j]

            # Use GetTFDBetweenConformers for same molecule different conformers
            tfd_values = TorsionFingerprints.GetTFDBetweenConformers(
                self.rdkit_mol,
                confIds1=[conf_id1],  # Single conformer list
                confIds2=[conf_id2],  # Single conformer list
                useWeights=self.use_weights,
                maxDev=self.max_dev,
                symmRadius=self.symm_radius,
                ignoreColinearBonds=self.ignore_colinear_bonds,
            )

            # GetTFDBetweenConformers returns a list, get the first (and only) value
            return tfd_values[0] if tfd_values else float("inf")

        except Exception as e:
            logger.warning(
                f"TFD calculation failed for conformers {i}, {j}: {str(e)}"
            )
            return float("inf")

    def _group_by_threshold(self, tfd_values, indices, n):
        """Threshold-based grouping for TFD using complete linkage."""
        # Build adjacency matrix for clustering (TFD uses <=)
        adj_matrix = np.zeros((n, n), dtype=bool)
        for (i, j), tfd in zip(indices, tfd_values):
            if tfd <= self.threshold:  # TFD: lower values = more similar
                adj_matrix[i, j] = adj_matrix[j, i] = True

        # Complete linkage grouping
        groups, index_groups = self._complete_linkage_grouping(adj_matrix, n)

        return groups, index_groups

    def _complete_linkage_grouping(self, adj_matrix, n):
        """
        Perform complete linkage grouping for TFD (from last to first).

        A structure joins a group only if its TFD to ALL existing members
        of that group is below the threshold.
        """
        assigned = [False] * n
        groups = []
        index_groups = []

        # Iterate from last to first (higher energy structures first as seeds)
        for i in range(n - 1, -1, -1):
            if assigned[i]:
                continue

            # Start a new group with molecule i
            current_group = [i]
            assigned[i] = True

            # Try to add unassigned molecules with lower indices (lower energy)
            for j in range(i - 1, -1, -1):
                if assigned[j]:
                    continue

                # Check if j is connected to ALL members in current_group
                can_join = all(
                    adj_matrix[j, member] for member in current_group
                )

                if can_join:
                    current_group.append(j)
                    assigned[j] = True

            # Sort indices within group (lowest index = lowest energy first)
            current_group.sort()
            # Add the completed group
            groups.append([self.molecules[idx] for idx in current_group])
            index_groups.append(current_group)

        # Reverse groups so that groups containing lower-energy structures come first
        groups.reverse()
        index_groups.reverse()

        return groups, index_groups

    def _group_by_num_groups(self, tfd_values, indices, n):
        """Automatic grouping to create specified number of groups for TFD."""
        if self.num_groups >= n:
            # If requesting more groups than molecules, each molecule is its own group
            print(
                f"[{self.__class__.__name__}] Requested {self.num_groups} groups but only {n} molecules. Creating {n} groups."
            )
            groups = [[mol] for mol in self.molecules]
            index_groups = [[i] for i in range(n)]
            return groups, index_groups

        # Find appropriate threshold to create desired number of groups
        threshold = self._find_optimal_tfd_threshold(tfd_values, indices, n)

        # Store the auto-determined threshold for summary reporting
        self._auto_threshold = threshold

        print(
            f"[{self.__class__.__name__}] Auto-determined threshold: {threshold:.7f} to create {self.num_groups} groups"
        )

        # Build adjacency matrix with the determined threshold
        adj_matrix = np.zeros((n, n), dtype=bool)
        for (i, j), tfd in zip(indices, tfd_values):
            if tfd <= threshold:  # TFD: lower values = more similar
                adj_matrix[i, j] = adj_matrix[j, i] = True

        # Complete linkage grouping
        groups, index_groups = self._complete_linkage_grouping(adj_matrix, n)
        actual_groups = len(groups)

        print(
            f"[{self.__class__.__name__}] Created {actual_groups} groups (requested: {self.num_groups})"
        )

        if actual_groups > self.num_groups:
            # If we have too many groups, merge the smallest ones
            groups, index_groups = self._merge_groups_to_target_tfd(
                groups, index_groups, adj_matrix
            )

        return groups, index_groups

    def _merge_groups_to_target_tfd(self, groups, index_groups, adj_matrix):
        """Merge groups to reach target number for TFD grouping."""
        while len(groups) > self.num_groups:
            min_size = float("inf")
            min_idx = 0
            for i, g in enumerate(groups):
                if len(g) < min_size:
                    min_size = len(g)
                    min_idx = i

            # Merge into largest group
            largest_idx = max(
                range(len(groups)),
                key=lambda i: len(groups[i]) if i != min_idx else -1,
            )
            groups[largest_idx].extend(groups[min_idx])
            index_groups[largest_idx].extend(index_groups[min_idx])

            groups.pop(min_idx)
            index_groups.pop(min_idx)

        return groups, index_groups

    def _find_optimal_tfd_threshold(self, tfd_values, indices, n):
        """Find threshold that creates approximately the desired number of groups using binary search."""
        # Sort TFD values (lower = more similar for TFD, so ascending order)
        sorted_tfd = sorted([tfd for tfd in tfd_values if not np.isinf(tfd)])

        if not sorted_tfd:
            return 0.0

        # Binary search for optimal threshold
        low, high = 0, len(sorted_tfd) - 1
        best_threshold = sorted_tfd[-1]

        while low <= high:
            mid = (low + high) // 2
            threshold = sorted_tfd[mid]

            # Build adjacency matrix with this threshold
            adj_matrix = np.zeros((n, n), dtype=bool)
            for (idx_i, idx_j), tfd in zip(indices, tfd_values):
                if tfd <= threshold:  # TFD: lower values = more similar
                    adj_matrix[idx_i, idx_j] = adj_matrix[idx_j, idx_i] = True

            # Count groups using complete linkage
            num_groups_found = self._count_complete_linkage_groups(
                adj_matrix, n
            )

            if num_groups_found == self.num_groups:
                # Found exact match
                return threshold
            elif num_groups_found > self.num_groups:
                # Too many groups, need higher threshold (more permissive)
                low = mid + 1
            else:
                # Too few groups, need lower threshold (more restrictive)
                best_threshold = threshold
                high = mid - 1

        return best_threshold

    def _count_complete_linkage_groups(self, adj_matrix, n):
        """Count number of groups using complete linkage algorithm."""
        assigned = [False] * n
        num_groups = 0

        for i in range(n):
            if assigned[i]:
                continue
            current_group = [i]
            assigned[i] = True

            for j in range(i + 1, n):
                if assigned[j]:
                    continue
                can_join = all(
                    adj_matrix[j, member] for member in current_group
                )
                if can_join:
                    current_group.append(j)
                    assigned[j] = True

            num_groups += 1

        return num_groups

    def _merge_smallest_tfd_groups(self, labels, actual_groups):
        """Merge smallest groups to reach target number of groups."""
        unique_labels = np.unique(labels)
        group_sizes = [
            (label, len(np.where(labels == label)[0]))
            for label in unique_labels
        ]
        group_sizes.sort(key=lambda x: x[1])  # Sort by size

        # Keep largest groups, merge smallest ones into the largest
        groups_to_keep = group_sizes[-self.num_groups :]
        groups_to_merge = group_sizes[: -self.num_groups]

        # Create final groups
        groups = []
        index_groups = []

        # Add the groups we're keeping
        for label, _ in groups_to_keep:
            indices = list(np.where(labels == label)[0])
            groups.append([self.molecules[i] for i in indices])
            index_groups.append(indices)

        # Merge remaining groups into the largest group
        if groups_to_merge:
            merge_indices = []
            for label, _ in groups_to_merge:
                merge_indices.extend(list(np.where(labels == label)[0]))

            # Add to the largest group
            if groups:
                groups[0].extend([self.molecules[i] for i in merge_indices])
                index_groups[0].extend(merge_indices)

        return groups, index_groups

    def group(self) -> Tuple[List[List[Molecule]], List[List[int]]]:
        """
        Group conformers based on TFD similarity using the same workflow as RMSDGrouper.

        Returns:
            Tuple[List[List[Molecule]], List[List[int]]]: Tuple containing:
                - List of molecule groups (each group is a list of molecules)
                - List of index groups (corresponding indices for each group)
        """
        import time

        # Record start time for grouping process
        grouping_start_time = time.time()

        n = len(self.molecules)

        if n == 0:
            return [], []

        if n == 1:
            return [self.molecules], [[0]]

        # Check if we have valid conformers
        if self.rdkit_mol is None or len(self.valid_conformer_ids) == 0:
            logger.warning(
                "No valid conformers found, each molecule becomes its own group"
            )
            return [[mol] for mol in self.molecules], [[i] for i in range(n)]

        # Generate conformer pairs for TFD calculation (same as RMSD workflow)
        indices = [(i, j) for i in range(n) for j in range(i + 1, n)]
        total_pairs = len(indices)

        print(
            f"[{self.__class__.__name__}] Starting TFD calculation for {n} conformers ({total_pairs} pairs)"
        )
        print(f"  - TFD threshold: {self.threshold}")
        print(f"  - Use weights: {self.use_weights}")
        print(f"  - Max deviation: {self.max_dev}")
        print(f"  - Symmetry radius: {self.symm_radius}")
        print(f"  - Ignore colinear bonds: {self.ignore_colinear_bonds}")

        # Calculate TFD values with real-time output (same as RMSD workflow)
        tfd_values = []
        for idx, (i, j) in enumerate(indices):
            tfd = self._calculate_tfd((i, j))
            tfd_values.append(tfd)
            print(
                f"The {idx+1}/{total_pairs} pair (conformer{i+1}, conformer{j+1}) calculation finished, TFD= {tfd:.7f}"
            )

        # Build full TFD matrix (same as RMSD workflow)
        tfd_matrix = np.zeros((n, n))
        for (i, j), tfd in zip(indices, tfd_values):
            tfd_matrix[i, j] = tfd_matrix[j, i] = tfd

        # Save TFD matrix (same as RMSD workflow)
        import os

        # Use label for folder name if provided
        if self.label:
            output_dir = f"{self.label}_group_result"
        else:
            output_dir = "group_result"
        os.makedirs(output_dir, exist_ok=True)

        # Create filename based on label and threshold or num_groups
        # Format: {label}_{GrouperClass}_T{threshold}.xlsx or {label}_{GrouperClass}_N{num_groups}.xlsx
        label_prefix = f"{self.label}_" if self.label else ""
        if self.num_groups is not None:
            matrix_filename = os.path.join(
                output_dir,
                f"{label_prefix}{self.__class__.__name__}_N{self.num_groups}.xlsx",
            )
        else:
            matrix_filename = os.path.join(
                output_dir,
                f"{label_prefix}{self.__class__.__name__}_T{self.threshold}.xlsx",
            )

        # Choose grouping strategy based on parameters (do this first to set _auto_threshold)
        if self.num_groups is not None:
            groups, index_groups = self._group_by_num_groups(
                tfd_values, indices, n
            )
        else:
            groups, index_groups = self._group_by_threshold(
                tfd_values, indices, n
            )

        # Calculate total grouping time
        grouping_end_time = time.time()
        grouping_time = grouping_end_time - grouping_start_time

        # Save TFD matrix (after grouping to include auto-determined threshold)
        self._save_tfd_matrix(tfd_matrix, matrix_filename, grouping_time)

        # Cache results
        self._cached_groups = groups
        self._cached_group_indices = index_groups

        print(
            f"[{self.__class__.__name__}] Found {len(groups)} groups using TFD"
        )

        return groups, index_groups

    def _save_tfd_matrix(
        self,
        tfd_matrix: np.ndarray,
        filename: str,
        grouping_time: float = None,
    ):
        """Save TFD matrix to Excel file with 7 decimal precision."""
        n = tfd_matrix.shape[0]

        # Change extension to .xlsx if it's .txt
        if filename.endswith(".txt"):
            filename = filename[:-4] + ".xlsx"
        elif not filename.endswith(".xlsx"):
            filename = filename + ".xlsx"

        # Create DataFrame with simple numeric indices as row and column labels
        row_labels = [str(i + 1) for i in range(n)]
        col_labels = [str(j + 1) for j in range(n)]

        # Replace inf with NaN for Excel (will show as blank)
        matrix_display = np.where(np.isinf(tfd_matrix), np.nan, tfd_matrix)
        df = pd.DataFrame(matrix_display, index=row_labels, columns=col_labels)

        # Create Excel writer with openpyxl engine
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            # Write matrix to sheet starting from row 12 to leave room for header info
            df.to_excel(
                writer,
                sheet_name="TFD_Matrix",
                startrow=12,
                float_format="%.7f",
            )

            # Get the worksheet to add header information
            worksheet = writer.sheets["TFD_Matrix"]

            # Add header information
            row = 1
            worksheet[f"A{row}"] = (
                f"Full TFD Matrix ({n}x{n}) - {self.__class__.__name__}"
            )
            row += 1
            worksheet[f"A{row}"] = (
                "Based on Schulz-Gasch et al., JCIM, 1499-1512 (2012)"
            )
            row += 1

            if hasattr(self, "num_groups") and self.num_groups is not None:
                worksheet[f"A{row}"] = (
                    f"Requested Groups (-N): {self.num_groups}"
                )
                row += 1
                if (
                    hasattr(self, "_auto_threshold")
                    and self._auto_threshold is not None
                ):
                    worksheet[f"A{row}"] = (
                        f"Auto-determined Threshold: {self._auto_threshold:.7f}"
                    )
                    row += 1
            else:
                worksheet[f"A{row}"] = f"Threshold: {self.threshold}"
                row += 1

            # TFD specific parameters
            worksheet[f"A{row}"] = (
                f"Use Weights: {getattr(self, 'use_weights', True)}"
            )
            row += 1
            worksheet[f"A{row}"] = (
                f"Max Deviation: {getattr(self, 'max_dev', 'equal')}"
            )
            row += 1
            worksheet[f"A{row}"] = (
                f"Symmetry Radius: {getattr(self, 'symm_radius', 2)}"
            )
            row += 1
            worksheet[f"A{row}"] = (
                f"Ignore Colinear Bonds: {getattr(self, 'ignore_colinear_bonds', True)}"
            )
            row += 1

            if grouping_time is not None:
                worksheet[f"A{row}"] = (
                    f"Grouping Time: {grouping_time:.2f} seconds"
                )
                row += 1

            worksheet[f"A{row}"] = (
                "Lower values indicate higher torsional similarity. Empty cells indicate calculation failures."
            )
            row += 1

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = min(max_length + 2, 18)
                worksheet.column_dimensions[column_letter].width = (
                    adjusted_width
                )

        logger.info(f"TFD matrix saved to {filename}")

    def calculate_full_tfd_matrix(
        self, output_file: Optional[str] = None
    ) -> np.ndarray:
        """
        Calculate full TFD matrix (same interface as RMSDGrouper).

        Args:
            output_file (Optional[str]): Path to save TFD matrix.

        Returns:
            np.ndarray: Symmetric TFD matrix (n x n).
        """
        n = len(self.molecules)

        if n == 0:
            return np.array([])

        if self.rdkit_mol is None or len(self.valid_conformer_ids) == 0:
            logger.warning(
                "No valid conformers, returning matrix of infinities"
            )
            return np.full((n, n), np.inf)

        logger.info(f"Calculating full TFD matrix for {n} conformers")

        # Use GetTFDMatrix for efficient calculation of all pairs
        try:
            # Get the lower triangular matrix from RDKit
            tfd_lower = TorsionFingerprints.GetTFDMatrix(
                self.rdkit_mol,
                useWeights=self.use_weights,
                maxDev=self.max_dev,
                symmRadius=self.symm_radius,
                ignoreColinearBonds=self.ignore_colinear_bonds,
            )

            # Reconstruct full symmetric matrix
            tfd_matrix = np.zeros((n, n))
            idx = 0
            for i in range(n):
                for j in range(i + 1, n):
                    if idx < len(tfd_lower):
                        tfd_matrix[i, j] = tfd_matrix[j, i] = tfd_lower[idx]
                        idx += 1

        except Exception as e:
            logger.warning(
                f"GetTFDMatrix failed: {e}, using pairwise calculation"
            )
            # Fallback to pairwise calculation
            tfd_matrix = np.zeros((n, n))
            for i in range(n):
                for j in range(i + 1, n):
                    tfd = self._calculate_tfd((i, j))
                    tfd_matrix[i, j] = tfd_matrix[j, i] = tfd

        # Output results (same format as RMSD)
        col_width = 12  # Fixed width: fits "-XX.XXXXXX" (6 decimals)
        row_header_width = 7  # Width for row headers

        print(f"\nFull TFD Matrix ({n}x{n}):")
        print("=" * (row_header_width + col_width * n))

        # Print column index header (right-aligned to match matrix values)
        header_line = " " * row_header_width
        for j in range(n):
            header_line += f"{j+1:>{col_width}}"
        print(header_line)

        # Print header row label with separator
        separator_line = "Conf".rjust(row_header_width)
        for j in range(n):
            separator_line += "-" * col_width
        print(separator_line)

        # Print matrix rows (right-aligned values)
        for i in range(n):
            row_line = f"{i+1:>{row_header_width}}"
            for j in range(n):
                if np.isinf(tfd_matrix[i, j]):
                    row_line += f"{'∞':>{col_width}}"
                else:
                    row_line += f"{tfd_matrix[i, j]:>{col_width}.6f}"
            print(row_line)

        # Save to file if requested
        if output_file:
            self._save_tfd_matrix(tfd_matrix, output_file)

        return tfd_matrix

    def __repr__(self):
        if self.num_groups is not None:
            return (
                f"{self.__class__.__name__}(num_groups={self.num_groups}, "
                f"num_procs={self.num_procs}, use_weights={self.use_weights}, "
                f"max_dev='{self.max_dev}', symm_radius={self.symm_radius})"
            )
        else:
            return (
                f"{self.__class__.__name__}(threshold={self.threshold}, "
                f"num_procs={self.num_procs}, use_weights={self.use_weights}, "
                f"max_dev='{self.max_dev}', symm_radius={self.symm_radius})"
            )


class StructureGrouperFactory:
    """
    Factory for creating molecular grouper instances.

    Provides a unified entry point to construct groupers by name. Supported
    strategies (case-insensitive):
    - "rmsd": BasicRMSDGrouper
    - "hrmsd": HungarianRMSDGrouper
    - "spyrmsd": SpyRMSDGrouper
    - "pymol" or "pymol_align": PymolRMSDGrouper
    - "tanimoto" or "fingerprint": TanimotoSimilarityGrouper
    - "torsion": TorsionFingerprintGrouper
    - "isomorphism" or "rdkit": RDKitIsomorphismGrouper
    - "formula": FormulaGrouper
    - "connectivity": ConnectivityGrouper

    Additional keyword arguments are forwarded to the specific grouper
    constructors (e.g., thresholds or flags).
    """

    pass


# Register grouper classes in GrouperJobRunner
GrouperJobRunner.GROUPER_CLASSES = {
    "rmsd": BasicRMSDGrouper,
    "hrmsd": HungarianRMSDGrouper,
    "spyrmsd": SpyRMSDGrouper,
    "irmsd": IRMSDGrouper,
    "pymolrmsd": PymolRMSDGrouper,
    "tanimoto": TanimotoSimilarityGrouper,
    "torsion": TorsionFingerprintGrouper,
    "isomorphism": RDKitIsomorphismGrouper,
    "formula": FormulaGrouper,
    "connectivity": ConnectivityGrouper,
}
